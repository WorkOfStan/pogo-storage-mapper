# pylint: disable=protected-access
from __future__ import annotations

import csv
import json
import os
import shutil
import threading
from collections import Counter
from pathlib import Path
from typing import cast

import pytest
from conftest import appraisal_values
from openpyxl import load_workbook  # pylint: disable=import-error
from PIL import Image

import pogo_storage_mapper.export as export_module
from pogo_storage_mapper.export import (
    EXPORT_COLUMNS,
    _scan_production_sequences,
    assemble_export_rows,
    run_production_export,
    write_export_csv,
    write_export_xlsx,
)
from pogo_storage_mapper.extract import (
    IV_NUMERIC_FIELD_NAMES,
    FragmentField,
    PokemonFragment,
)
from pogo_storage_mapper.scan_frames import (
    FEATURE_KEYS,
    NON_EXTRACTABLE_CLASS,
    FrameCandidate,
    FrameScanRecord,
    FrameVisualRecord,
    ProductionSequenceScanResult,
    ScanSettings,
    SourceAsset,
    _empty_iv_evidence_from_signals,
    _IvEvidence,
)


def _field(value: str | int | float | bool) -> FragmentField:
    return FragmentField(value=value, source="test", evidence="test")


def _fragment(
    fields: dict[str, str | int | float | bool],
    *,
    frame_index: int = 0,
    fragment_type: str = "detail",
) -> PokemonFragment:
    return PokemonFragment(
        source_file="source.mp4",
        source_type="video",
        frame_index=frame_index,
        timestamp_s=float(frame_index),
        classification="detail",
        raw_classification=("appraisal" if fragment_type == "appraisal" else "detail"),
        fragment_type=fragment_type,
        fields={name: _field(value) for name, value in fields.items()},
    )


def _core_fields(
    **extra: str | int | float | bool,
) -> dict[str, str | int | float | bool]:
    return {
        "hp_current": 10,
        "hp_max": 10,
        "weight_kg": 1.25,
        **extra,
    }


def _detail_anchor_fields(
    **extra: str | int | float | bool,
) -> dict[str, str | int | float | bool]:
    return _core_fields(
        fast_move_key="vine-whip-fast",
        fast_move_name="Vine Whip",
        charged_move_key="solar-beam",
        charged_move_name="Solar Beam",
        **extra,
    )


def _row_by_species(
    rows: list[dict[str, export_module.ExportValue]], species_key: str
) -> dict[str, export_module.ExportValue]:
    return next(row for row in rows if row["species_key"] == species_key)


def test_assemble_export_rows_accepts_core_identity_with_cp() -> None:
    result = assemble_export_rows(
        [[_fragment(_detail_anchor_fields(cp=321, display_name_text="Buddy"))]]
    )

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["source_file"] == "source.mp4"
    assert row["hp_current"] == 10
    assert row["weight_kg"] == 1.25
    assert row["cp"] == 321
    assert row["display_name"] == "Buddy"
    assert result.rejected_sequence_count == 0


def test_assemble_export_rows_merges_detail_appraisal_story_and_iv() -> None:
    detail = _fragment(
        _core_fields(
            canonical_name_text="Bulbasaur",
            catch_date_text="1/2/2026",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            species_key="bulbasaur",
            species_name="Bulbasaur",
            pokedex_id=1,
        ),
        frame_index=3,
    )
    appraisal_fields = cast(
        dict[str, str | int | float | bool],
        {"iv_complete": True, "iv_star_agreement": True, **appraisal_values()},
    )
    appraisal = _fragment(
        appraisal_fields,
        frame_index=4,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail, appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["canonical_name"] == "Bulbasaur"
    assert row["species_key"] == "bulbasaur"
    assert row["catch_location"] == "Prague, Czechia"
    assert row["iv_attack"] == 13
    assert row["iv_complete"] is True
    assert row["first_frame_index"] == 3
    assert row["last_frame_index"] == 4


def test_assemble_export_rows_rejects_conflicting_evidence() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(_core_fields(cp=100, fast_move_key="vine-whip")),
                _fragment(
                    {
                        "iv_attack": 14,
                        "iv_defense": 15,
                        "iv_stamina": 15,
                        "canonical_name_text": "Bulbasaur",
                    },
                    frame_index=1,
                ),
                _fragment(
                    {
                        "cp": 200,
                        "iv_attack": 15,
                        "iv_defense": 15,
                        "iv_stamina": 14,
                        "canonical_name_text": "Charmander",
                        "fast_move_key": "ember",
                    },
                    frame_index=2,
                ),
            ]
        ]
    )

    assert not result.rows
    messages = [warning.message for warning in result.warnings]
    assert any("cp" in message for message in messages)
    assert any("iv_attack" in message for message in messages)
    assert any("canonical_name" in message for message in messages)
    assert any("fast_move_key" in message for message in messages)


def test_assemble_export_rows_merges_same_core_identity() -> None:
    result = assemble_export_rows(
        [
            [_fragment(_core_fields(cp=100), frame_index=1)],
            [
                _fragment(
                    _detail_anchor_fields(
                        canonical_name_text="Bulbasaur",
                        catch_date_text="1/2/2026",
                        location_text="Prague, Czechia",
                    ),
                    frame_index=20,
                )
            ],
        ]
    )

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 100
    assert row["canonical_name"] == "Bulbasaur"
    assert row["first_frame_index"] == 1
    assert row["last_frame_index"] == 20
    assert result.rejected_sequence_count == 1


def test_assemble_export_rows_merges_physical_support_into_anchored_identity() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(
                    _core_fields(
                        cp=1206,
                        canonical_name_text="Gengar",
                        species_key="gengar",
                        species_name="Gengar",
                        pokedex_id=94,
                        catch_date_text="12/7/2024",
                        location_text="Prague, Czechia",
                        iv_attack=13,
                        iv_defense=11,
                        iv_stamina=14,
                    ),
                    frame_index=496,
                    fragment_type="appraisal",
                )
            ],
            [
                _fragment(
                    _core_fields(
                        fast_move_key="sucker-punch-fast",
                        fast_move_name="Sucker Punch",
                        charged_move_key="sludge-bomb",
                        charged_move_name="Sludge Bomb",
                    ),
                    frame_index=282,
                )
            ],
            [_fragment(_core_fields(), frame_index=281)],
        ]
    )

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["canonical_name"] == "Gengar"
    assert row["fast_move_name"] == "Sucker Punch"
    assert row["charged_move_name"] == "Sludge Bomb"
    assert row["first_frame_index"] == 281
    assert row["last_frame_index"] == 496
    assert result.rejected_sequence_count == 2


def test_assemble_export_rows_keeps_hp_weight_without_moves_as_support_only() -> None:
    result = assemble_export_rows([[_fragment(_core_fields(cp=321), frame_index=10)]])

    assert not result.rows
    assert result.rejected_sequence_count == 1
    assert any("support-only" in warning.message for warning in result.warnings)
    assert result.unresolved_pokemon_like_sequence_count == 1


def test_assemble_export_rows_accepts_complementary_partial_identity() -> None:
    detail_support = _fragment(
        _core_fields(
            cp=1397,
            hp_current=87,
            hp_max=87,
            weight_kg=25.67,
            height_m=0.85,
        ),
        frame_index=137,
    )
    appraisal_support = _fragment(
        {
            "hp_current": 87,
            "hp_max": 87,
            "species_key": "chandelure",
            "species_name": "Chandelure",
            "pokedex_id": 609,
            "canonical_name_text": "Chandelure",
            "catch_date_text": "10/21/2023",
            "location_text": "Prague, Czechia",
            "catch_country_text": "Czechia",
            "iv_complete": True,
            "iv_star_agreement": True,
            "iv_attack": 15,
            "iv_defense": 14,
            "iv_stamina": 15,
            "iv_sum": 44,
            "appraisal_star_count": 3,
        },
        frame_index=874,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail_support], [appraisal_support]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["canonical_name"] == "Chandelure"
    assert row["cp"] == 1397
    assert row["hp_current"] == 87
    assert row["weight_kg"] == 25.67
    assert row["iv_attack"] == 15
    assert any(
        "Accepted partial export row" in warning.message for warning in result.warnings
    )
    assert result.unresolved_pokemon_like_sequence_count == 0


def test_assemble_export_rows_prefers_appraisal_species_for_same_physical_key() -> None:
    chandelure = _fragment(
        _core_fields(
            hp_current=87,
            hp_max=87,
            weight_kg=25.67,
            cp=1397,
            species_key="chandelure",
            species_name="Chandelure",
            pokedex_id=609,
            canonical_name_text="Chandelure",
            catch_date_text="10/21/2023",
            catch_country_text="Czechia",
            iv_attack=15,
            iv_defense=14,
            iv_stamina=15,
        ),
        frame_index=243,
        fragment_type="appraisal",
    )
    litwick = _fragment(
        _detail_anchor_fields(
            hp_current=87,
            hp_max=87,
            weight_kg=25.67,
            species_key="litwick",
            species_name="Litwick",
            pokedex_id=607,
            canonical_name_text="Litwick",
        ),
        frame_index=2009,
    )

    result = assemble_export_rows([[chandelure], [litwick]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "chandelure"
    assert row["canonical_name"] == "Chandelure"
    assert row["fast_move_name"] == "Vine Whip"
    assert not any(
        "Rejected conflicting export evidence for species_key" in warning.message
        for warning in result.warnings
    )
    assert any(
        "Ignored conflicting detail identity evidence for species_key"
        in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_pairs_detail_and_appraisal_anchors() -> None:
    detail = _fragment(
        _detail_anchor_fields(cp=936, height_m=1.17),
        frame_index=610,
    )
    appraisal = _fragment(
        _core_fields(
            species_key="ivysaur",
            species_name="Ivysaur",
            pokedex_id=2,
            canonical_name_text="Ivysaur",
            catch_date_text="1/30/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=11,
            iv_defense=14,
            iv_stamina=12,
            iv_sum=37,
            appraisal_star_count=3,
            iv_star_agreement=True,
        ),
        frame_index=150,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 936
    assert row["species_key"] == "ivysaur"
    assert row["iv_attack"] == 11
    assert row["fast_move_key"] == "vine-whip-fast"


def test_assemble_export_rows_keeps_stuttered_cp3233_appraisal() -> None:
    detail = _fragment(
        _detail_anchor_fields(
            hp_current=180,
            hp_max=180,
            weight_kg=162.59,
            height_m=1.62,
        ),
        frame_index=118,
    )
    noisy_appraisal = _fragment(
        _core_fields(
            hp_current=180,
            hp_max=180,
            weight_kg=162.59,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="8/27/2022",
            location_text="Jihocesky kraj, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=15,
            iv_stamina=15,
            iv_sum=45,
            appraisal_star_count=4,
            appraisal_perfect=True,
        ),
        frame_index=3688,
        fragment_type="appraisal",
    )
    stutter_appraisal = _fragment(
        _core_fields(
            hp_current=180,
            hp_max=180,
            weight_kg=162.59,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="8/27/2022",
            location_text="Jihoc\u00e9esky kraj, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=15,
            iv_stamina=15,
            iv_sum=45,
            appraisal_star_count=4,
            appraisal_perfect=True,
        ),
        frame_index=3723,
        fragment_type="appraisal",
    )
    cp_appraisal = _fragment(
        _core_fields(
            hp_current=180,
            hp_max=180,
            weight_kg=162.59,
            cp=3233,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="8/27/2022",
            location_text="Jihoc\u00e9esky kraj, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=15,
            iv_stamina=15,
            iv_sum=45,
            appraisal_star_count=4,
            appraisal_perfect=True,
        ),
        frame_index=3840,
        fragment_type="appraisal",
    )

    result = assemble_export_rows(
        [[detail], [noisy_appraisal], [stutter_appraisal], [cp_appraisal]]
    )

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 3233
    assert row["species_key"] == "machamp"
    assert row["hp_current"] == 180
    assert row["weight_kg"] == 162.59
    assert row["catch_location"] == "Jihoc\u00e9esky kraj, Czechia"
    assert row["fast_move_key"] == "vine-whip-fast"
    assert row["charged_move_key"] == "solar-beam"
    assert row["iv_complete"] is True
    assert row["appraisal_perfect"] is True
    assert any(
        "Ignored conflicting export evidence for catch_location after consensus"
        in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_merges_move_detail_into_appraisal_row() -> None:
    detail = _fragment(
        _detail_anchor_fields(hp_current=132, hp_max=132, weight_kg=135.9),
        frame_index=564,
    )
    appraisal = _fragment(
        _core_fields(
            hp_current=132,
            hp_max=132,
            weight_kg=135.9,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="6/6/2025",
            location_text="Hlavni mesto Praha, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=12,
            iv_stamina=15,
            iv_sum=42,
            appraisal_star_count=3,
        ),
        frame_index=3240,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[appraisal], [detail]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "machamp"
    assert row["fast_move_name"] == "Vine Whip"
    assert row["charged_move_name"] == "Solar Beam"
    assert row["iv_attack"] == 15


def test_assemble_export_rows_recovers_same_hp_weight_mismatch_same_species() -> None:
    detail = _fragment(
        _detail_anchor_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=66.05,
            species_key="enamorus",
            species_name="Enamorus",
            pokedex_id=905,
            canonical_name_text="Enamorus",
        ),
        frame_index=3084,
    )
    appraisal = _fragment(
        _core_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=6.05,
            species_key="enamorus",
            species_name="Enamorus",
            pokedex_id=905,
            canonical_name_text="Enamorus",
            catch_date_text="2/8/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=13,
            iv_stamina=13,
            iv_sum=41,
            appraisal_star_count=3,
        ),
        frame_index=479,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "enamorus"
    assert row["weight_kg"] == 6.05
    assert row["fast_move_key"] == "vine-whip-fast"
    assert row["iv_attack"] == 15
    assert any(
        "Recovered detail/appraisal match despite physical-key mismatch"
        in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_recovers_unique_detail_without_species() -> None:
    detail = _fragment(
        _detail_anchor_fields(
            hp_current=87,
            hp_max=87,
            weight_kg=26.57,
        ),
        frame_index=137,
    )
    appraisal = _fragment(
        _core_fields(
            hp_current=87,
            hp_max=87,
            weight_kg=25.67,
            species_key="chandelure",
            species_name="Chandelure",
            pokedex_id=609,
            canonical_name_text="Chandelure",
            catch_date_text="10/21/2023",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=14,
            iv_stamina=15,
            iv_sum=44,
            appraisal_star_count=3,
        ),
        frame_index=874,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "chandelure"
    assert row["fast_move_name"] == "Vine Whip"
    assert row["iv_sum"] == 44


def test_assemble_export_rows_does_not_recover_ambiguous_same_hp_details() -> None:
    appraisal = _fragment(
        _core_fields(
            hp_current=132,
            hp_max=132,
            weight_kg=135.9,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="6/6/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=12,
            iv_stamina=15,
            iv_sum=42,
            appraisal_star_count=3,
        ),
        frame_index=3240,
        fragment_type="appraisal",
    )
    first_detail = _fragment(
        _detail_anchor_fields(hp_current=132, hp_max=132, weight_kg=136.9),
        frame_index=564,
    )
    second_detail = _fragment(
        _detail_anchor_fields(hp_current=132, hp_max=132, weight_kg=137.9),
        frame_index=565,
    )

    result = assemble_export_rows([[appraisal], [first_detail], [second_detail]])

    assert len(result.rows) == 3
    assert not any(
        "Recovered detail/appraisal match despite physical-key mismatch"
        in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_recovers_unique_species_pair_from_multiple() -> None:
    enamorus_appraisal = _fragment(
        _core_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=6.05,
            species_key="enamorus",
            species_name="Enamorus",
            pokedex_id=905,
            canonical_name_text="Enamorus",
            catch_date_text="2/8/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=13,
            iv_stamina=13,
            iv_sum=41,
            appraisal_star_count=3,
        ),
        frame_index=479,
        fragment_type="appraisal",
    )
    arcanine_appraisal = _fragment(
        _core_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=177.07,
            species_key="arcanine",
            species_name="Arcanine",
            pokedex_id=59,
            canonical_name_text="Arcanine",
            catch_date_text="8/26/2023",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=0,
            iv_stamina=15,
            iv_sum=30,
            appraisal_star_count=2,
        ),
        frame_index=617,
        fragment_type="appraisal",
    )
    detail = _fragment(
        _detail_anchor_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=66.05,
            species_key="enamorus",
            species_name="Enamorus",
            pokedex_id=905,
            canonical_name_text="Enamorus",
        ),
        frame_index=3084,
    )

    result = assemble_export_rows(
        [[enamorus_appraisal], [arcanine_appraisal], [detail]]
    )

    assert len(result.rows) == 2
    enamorus = _row_by_species(result.rows, "enamorus")
    arcanine = _row_by_species(result.rows, "arcanine")
    assert enamorus["fast_move_key"] == "vine-whip-fast"
    assert enamorus["iv_attack"] == 15
    assert arcanine["fast_move_key"] is None


def test_assemble_export_rows_blocks_recovery_on_species_conflict() -> None:
    appraisal = _fragment(
        _core_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=6.05,
            species_key="enamorus",
            species_name="Enamorus",
            pokedex_id=905,
            canonical_name_text="Enamorus",
            catch_date_text="2/8/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=13,
            iv_stamina=13,
            iv_sum=41,
            appraisal_star_count=3,
        ),
        frame_index=479,
        fragment_type="appraisal",
    )
    detail = _fragment(
        _detail_anchor_fields(
            hp_current=114,
            hp_max=114,
            weight_kg=66.05,
            species_key="arcanine",
            species_name="Arcanine",
            pokedex_id=59,
            canonical_name_text="Arcanine",
        ),
        frame_index=3084,
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 2
    assert not any(
        "Recovered detail/appraisal match despite physical-key mismatch"
        in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_logs_missing_moves_without_detail_candidate() -> None:
    appraisal = _fragment(
        _core_fields(
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="6/6/2025",
            location_text="Hlavni mesto Praha, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=12,
            iv_stamina=15,
            iv_sum=42,
            appraisal_star_count=3,
        ),
        frame_index=3240,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[appraisal]])

    assert len(result.rows) == 1
    assert result.rows[0]["fast_move_key"] is None
    assert any(
        "missing moves: no candidate" in warning.message for warning in result.warnings
    )
    assert result.row_diagnostics[0]["moves_status"] == "appraisal_without_detail_moves"
    assert result.row_diagnostics[0]["source_appraisal_frames"] == [3240]


def test_assemble_export_rows_logs_missing_moves_for_detail_without_moves() -> None:
    detail = _fragment(
        _core_fields(hp_current=132, hp_max=132, weight_kg=135.9),
        frame_index=564,
    )
    appraisal = _fragment(
        _core_fields(
            hp_current=132,
            hp_max=132,
            weight_kg=135.9,
            species_key="machamp",
            species_name="Machamp",
            pokedex_id=68,
            canonical_name_text="Machamp",
            catch_date_text="6/6/2025",
            location_text="Hlavni mesto Praha, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_star_agreement=True,
            iv_attack=15,
            iv_defense=12,
            iv_stamina=15,
            iv_sum=42,
            appraisal_star_count=3,
        ),
        frame_index=3240,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[appraisal], [detail]])

    assert len(result.rows) == 1
    assert result.rows[0]["fast_move_key"] is None
    assert any(
        "missing moves: rejected candidate" in warning.message
        for warning in result.warnings
    )
    detail_diagnostics = [
        diagnostic
        for diagnostic in result.row_diagnostics
        if diagnostic["source_detail_frames"] == [564]
    ]
    assert detail_diagnostics[0]["moves_status"] == "detail_without_resolved_moves"


def test_assemble_export_rows_prefers_detail_cp_over_low_appraisal_noise() -> None:
    detail = _fragment(
        _detail_anchor_fields(cp=111, height_m=1.54),
        frame_index=372,
    )
    appraisal = _fragment(
        _core_fields(
            cp=96,
            species_key="scyther",
            species_name="Scyther",
            pokedex_id=123,
            canonical_name_text="Scyther",
            catch_date_text="12/2/2024",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=False,
            iv_attack=7,
            iv_defense=9,
            iv_stamina=6,
            iv_sum=22,
            appraisal_star_count=0,
        ),
        frame_index=463,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 111
    assert row["fast_move_key"] == "vine-whip-fast"
    assert row["species_key"] == "scyther"
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_keeps_identity_when_weak_merge_fields_conflict() -> None:
    detail = _fragment(
        _detail_anchor_fields(
            cp=936,
            height_m=1.17,
            display_name_text="Ivysaur",
        ),
        frame_index=610,
    )
    appraisal = _fragment(
        _core_fields(
            height_m=4.0,
            display_name_text="lvysaur q",
            species_key="ivysaur",
            species_name="Ivysaur",
            pokedex_id=2,
            canonical_name_text="Ivysaur",
            catch_date_text="1/30/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=11,
            iv_defense=14,
            iv_stamina=12,
            iv_sum=37,
            appraisal_star_count=3,
            iv_star_agreement=True,
        ),
        frame_index=154,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[detail], [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "ivysaur"
    assert row["fast_move_key"] == "vine-whip-fast"
    assert row["display_name"] is None
    assert row["height_m"] == 1.17
    assert result.rejected_sequence_count == 1
    assert any(
        "Ignored conflicting weak export evidence for display_name" in warning.message
        for warning in result.warnings
    )
    assert any(
        "Ignored conflicting weak export evidence for height_m" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_support_conflicts_do_not_veto_anchor() -> None:
    result = assemble_export_rows(
        [
            [_fragment(_detail_anchor_fields(cp=100), frame_index=20)],
            [_fragment(_core_fields(cp=200), frame_index=10)],
        ]
    )

    assert len(result.rows) == 1
    assert result.rows[0]["cp"] is None
    assert result.rejected_sequence_count == 1
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_prefers_detail_height_over_appraisal_support() -> None:
    appraisal = _fragment(
        {
            **_core_fields(
                height_m=4.0,
                species_key="scyther",
                species_name="Scyther",
                pokedex_id=123,
                canonical_name_text="Scyther",
                catch_date_text="12/2/2024",
                location_text="Prague, Czechia",
                catch_country_text="Czechia",
                iv_complete=False,
                iv_attack=7,
                iv_defense=9,
                iv_stamina=6,
                iv_sum=22,
                appraisal_star_count=0,
                iv_star_agreement=False,
            )
        },
        frame_index=463,
        fragment_type="appraisal",
    )
    detail_support = _fragment(
        _core_fields(height_m=1.54),
        frame_index=380,
    )

    result = assemble_export_rows([[appraisal], [detail_support]])

    assert len(result.rows) == 1
    assert result.rows[0]["height_m"] == 1.54
    assert result.rows[0]["iv_complete"] is False
    assert any(
        "Accepted appraisal anchor without IV/star agreement" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_accepts_appraisal_without_star_agreement() -> None:
    appraisal = _fragment(
        cast(
            dict[str, str | int | float | bool],
            {
                **_core_fields(
                    species_key="toxel",
                    species_name="Toxel",
                    pokedex_id=848,
                    canonical_name_text="Toxel",
                    catch_date_text="12/11/2024",
                    location_text="Prague, Czechia",
                    catch_country_text="Czechia",
                    iv_complete=False,
                    iv_star_agreement=False,
                ),
                **appraisal_values(
                    iv_attack=14,
                    iv_defense=12,
                    iv_stamina=12,
                    iv_sum=38,
                    appraisal_star_count=2,
                ),
            },
        ),
        frame_index=618,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([[appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "toxel"
    assert row["iv_attack"] == 14
    assert row["iv_complete"] is False
    assert any(
        "Accepted appraisal anchor without IV/star agreement" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_blanks_conflicting_cp_for_same_core_identity() -> None:
    result = assemble_export_rows(
        [
            [_fragment(_detail_anchor_fields(cp=100), frame_index=1)],
            [_fragment(_detail_anchor_fields(cp=200), frame_index=20)],
        ]
    )

    assert len(result.rows) == 1
    assert result.rows[0]["cp"] is None
    assert result.rejected_sequence_count == 1
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_prefers_clean_cp_over_suffix_pollution() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(_core_fields(cp=589), frame_index=1),
                _fragment(_core_fields(cp=589), frame_index=2),
                _fragment(_core_fields(cp=589), frame_index=3),
                _fragment(_core_fields(cp=5891), frame_index=4),
                _fragment(_core_fields(cp=5891), frame_index=5),
                _fragment(_core_fields(cp=5891), frame_index=6),
                _fragment(_core_fields(cp=5891), frame_index=7),
                _fragment(_core_fields(cp=5894), frame_index=8),
                _fragment(_core_fields(cp=5894), frame_index=9),
                _fragment(_core_fields(cp=5894), frame_index=10),
            ]
        ]
    )

    assert not result.rows
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_keeps_clean_anchor_cp_over_suffix_pollution() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(_detail_anchor_fields(cp=464), frame_index=1),
                _fragment(_detail_anchor_fields(cp=464), frame_index=2),
                _fragment(_detail_anchor_fields(cp=464), frame_index=3),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=4),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=5),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=6),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=7),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=8),
                _fragment(_detail_anchor_fields(cp=4644), frame_index=9),
            ]
        ]
    )

    assert len(result.rows) == 1
    assert result.rows[0]["cp"] == 464
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_uses_same_hp_support_moves_and_clean_low_cp() -> None:
    appraisal = _fragment(
        _core_fields(
            cp=68,
            species_key="scyther",
            species_name="Scyther",
            pokedex_id=123,
            canonical_name_text="Scyther",
            catch_date_text="12/2/2024",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=7,
            iv_defense=9,
            iv_stamina=6,
            iv_sum=22,
            appraisal_star_count=0,
            iv_star_agreement=True,
        ),
        frame_index=463,
        fragment_type="appraisal",
    )
    moves = _fragment(
        _core_fields(
            cp=565,
            weight_kg=21.87,
            fast_move_key="air-slash-fast",
            fast_move_name="Air Slash",
            charged_move_key="frustration",
            charged_move_name="Frustration",
        ),
        frame_index=372,
    )
    conflicting_weight = _fragment(
        _core_fields(weight_kg=31.87),
        frame_index=371,
    )

    result = assemble_export_rows([[appraisal], [moves, conflicting_weight]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 565
    assert row["fast_move_name"] == "Air Slash"
    assert row["charged_move_name"] == "Frustration"
    assert row["is_shadow"] is True


def test_merge_uses_raw_cp_evidence_when_candidate_cp_was_unresolved() -> None:
    accepted_row = export_module._empty_row()
    accepted_row.update(
        {
            "source_file": "source.mp4",
            "source_type": "video",
            "first_frame_index": 441,
            "last_frame_index": 463,
            "first_timestamp_s": 0.0,
            "last_timestamp_s": 1.0,
            "species_key": "scyther",
            "species_name": "Scyther",
            "pokedex_id": 123,
            "canonical_name": "Scyther",
            "catch_date": "12/2/2024",
            "catch_location": "Prague, Czechia",
            "catch_country": "Czechia",
            "hp_current": 66,
            "hp_max": 66,
            "weight_kg": 51.87,
            "iv_complete": True,
            "iv_attack": 7,
            "iv_defense": 9,
            "iv_stamina": 6,
        }
    )
    support_row = export_module._empty_row()
    support_row.update(
        {
            "source_file": "source.mp4",
            "source_type": "video",
            "first_frame_index": 380,
            "last_frame_index": 413,
            "first_timestamp_s": 0.0,
            "last_timestamp_s": 1.0,
            "hp_current": 66,
            "hp_max": 66,
            "weight_kg": 51.87,
            "height_m": 1.54,
            "fast_move_key": "air-slash-fast",
            "fast_move_name": "Air Slash",
            "charged_move_key": "shadow-bone",
            "charged_move_name": "Shadow Bone",
        }
    )
    accepted = export_module._RowCandidate(
        accepted_row,
        ("physical", "source.mp4", "66", "66", "51.87"),
        accepted=True,
        anchor_kind="appraisal",
        scan_start_frame_index=441,
        fragment_types={"appraisal"},
        column_value_counts={"cp": Counter({68: 2, 96: 1, 5637: 1})},
    )
    support = export_module._RowCandidate(
        support_row,
        ("physical", "source.mp4", "66", "66", "51.87"),
        anchor_kind="support",
        scan_start_frame_index=380,
        fragment_types={"detail"},
        column_value_counts={"cp": Counter({5655: 2, 5653: 2, 5657: 2, 5654: 1})},
    )

    merged = export_module._merge_identity_candidates(
        ("physical", "source.mp4", "66", "66", "51.87"),
        [accepted, support],
    )

    assert merged.row["cp"] == 565
    assert merged.row["fast_move_name"] == "Air Slash"


def test_assemble_export_rows_preserves_clean_detail_and_appraisal_values() -> None:
    detail = _fragment(
        _core_fields(
            cp=399,
            fast_move_key="acid-fast",
            fast_move_name="Acid",
            charged_move_key="power-up-punch",
            charged_move_name="Power Up Punch",
        ),
        frame_index=90,
    )
    appraisal_early = _fragment(
        _core_fields(
            cp=399,
            species_key="toxel",
            species_name="Toxel",
            pokedex_id=848,
            canonical_name_text="Toxel",
            catch_date_text="12/11/2024",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=14,
            iv_defense=12,
            iv_stamina=12,
            iv_sum=38,
            appraisal_star_count=1,
            iv_star_agreement=True,
        ),
        frame_index=596,
        fragment_type="appraisal",
    )
    appraisal_clean = _fragment(
        _core_fields(
            cp=399,
            species_key="toxel",
            species_name="Toxel",
            pokedex_id=848,
            canonical_name_text="Toxel",
            catch_date_text="12/11/2024",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=14,
            iv_defense=12,
            iv_stamina=12,
            iv_sum=38,
            appraisal_star_count=3,
            iv_star_agreement=True,
        ),
        frame_index=604,
        fragment_type="appraisal",
    )
    noisy_support = _fragment(_core_fields(cp=329), frame_index=696)

    result = assemble_export_rows(
        [[detail], [appraisal_early, appraisal_clean], [noisy_support]]
    )

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["cp"] == 399
    assert row["appraisal_star_count"] == 3
    assert row["fast_move_name"] == "Acid"
    assert row["charged_move_name"] == "Power Up Punch"


def test_assemble_export_rows_keeps_stable_toxel_cp_over_noisy_later_cp() -> None:
    detail_frames = [
        _fragment(
            _core_fields(
                hp_current=78,
                hp_max=78,
                weight_kg=5.43,
                cp=399,
                fast_move_key="acid-fast",
                fast_move_name="Acid",
                charged_move_key="power-up-punch",
                charged_move_name="Power Up Punch",
            ),
            frame_index=frame_index,
        )
        for frame_index in (56, 57, 58)
    ]
    appraisal = _fragment(
        _core_fields(
            hp_current=78,
            hp_max=78,
            weight_kg=5.43,
            cp=329,
            species_key="toxel",
            species_name="Toxel",
            pokedex_id=848,
            canonical_name_text="Toxel",
            catch_date_text="12/11/2024",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=14,
            iv_defense=12,
            iv_stamina=12,
            iv_sum=38,
            appraisal_star_count=3,
            iv_star_agreement=True,
        ),
        frame_index=696,
        fragment_type="appraisal",
    )

    result = assemble_export_rows([detail_frames, [appraisal]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "toxel"
    assert row["hp_current"] == 78
    assert row["cp"] == 399
    assert row["charged_move_name"] == "Power Up Punch"


def test_assemble_export_rows_keeps_complete_iv_triplet_over_noisy_neighbors() -> None:
    fragments = [
        _fragment(
            _core_fields(
                cp=1446,
                species_key="toxtricity",
                species_name="Toxtricity",
                pokedex_id=849,
                canonical_name_text="Toxtricity",
                catch_date_text="8/24/2025",
                location_text="Prague, Czechia",
                catch_country_text="Czechia",
                iv_complete=True,
                iv_attack=14,
                iv_defense=11,
                iv_stamina=11,
                iv_sum=36,
                appraisal_star_count=2,
                appraisal_perfect=False,
                iv_star_agreement=True,
            ),
            frame_index=365,
            fragment_type="appraisal",
        ),
        _fragment(
            _core_fields(
                cp=1446,
                species_key="toxtricity",
                species_name="Toxtricity",
                pokedex_id=849,
                canonical_name_text="Toxtricity",
                catch_date_text="8/24/2025",
                location_text="Prague, Czechia",
                catch_country_text="Czechia",
                iv_complete=False,
                iv_attack=2,
                iv_defense=4,
                iv_stamina=11,
                iv_sum=17,
                appraisal_star_count=2,
                appraisal_perfect=False,
            ),
            frame_index=364,
            fragment_type="appraisal",
        ),
    ]

    result = assemble_export_rows([fragments])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["iv_complete"] is True
    assert row["iv_attack"] == 14
    assert row["iv_defense"] == 11
    assert row["iv_stamina"] == 11
    assert row["iv_sum"] == 36
    assert row["appraisal_star_count"] == 2
    assert row["appraisal_perfect"] is False
    messages = [warning.message for warning in result.warnings]
    assert any("complete same-frame IV" in message for message in messages)
    assert not any(
        "Rejected conflicting export evidence for iv_attack" in message
        for message in messages
    )


def test_assemble_export_rows_keeps_sinistea_cp_without_trailing_digit() -> None:
    fragments = [
        _fragment(
            _core_fields(
                hp_current=69,
                hp_max=69,
                weight_kg=0.29,
                cp=cp,
                species_key="sinistea",
                species_name="Sinistea",
                pokedex_id=854,
                canonical_name_text="Sinistea",
                catch_date_text="12/7/2024",
                location_text="Prague, Czechia",
                catch_country_text="Czechia",
                iv_complete=True,
                iv_attack=9,
                iv_defense=14,
                iv_stamina=14,
                iv_sum=37,
                appraisal_star_count=3,
                iv_star_agreement=True,
            ),
            frame_index=frame_index,
            fragment_type="appraisal",
        )
        for frame_index, cp in enumerate(
            (464, 464, 464, 4644, 4644, 4644, 4644, 4644),
            start=152,
        )
    ]
    moves = _fragment(
        _core_fields(
            hp_current=69,
            hp_max=69,
            weight_kg=0.29,
            cp=464,
            fast_move_key="astonish-fast",
            fast_move_name="Astonish",
            charged_move_key="shadow-ball",
            charged_move_name="Shadow Ball",
        ),
        frame_index=180,
    )

    result = assemble_export_rows([fragments, [moves]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "sinistea"
    assert row["hp_current"] == 69
    assert row["cp"] == 464
    assert row["charged_move_name"] == "Shadow Ball"


def test_assemble_export_rows_keeps_ivysaur_solar_beam_after_merge() -> None:
    appraisal = _fragment(
        _core_fields(
            hp_current=50,
            hp_max=99,
            weight_kg=16.08,
            cp=936,
            species_key="ivysaur",
            species_name="Ivysaur",
            pokedex_id=2,
            canonical_name_text="Ivysaur",
            catch_date_text="1/30/2025",
            location_text="Prague, Czechia",
            catch_country_text="Czechia",
            iv_complete=True,
            iv_attack=11,
            iv_defense=14,
            iv_stamina=12,
            iv_sum=37,
            appraisal_star_count=3,
            iv_star_agreement=True,
        ),
        frame_index=86,
        fragment_type="appraisal",
    )
    moves = _fragment(
        _core_fields(
            hp_current=50,
            hp_max=99,
            weight_kg=16.08,
            cp=936,
            fast_move_key="vine-whip-fast",
            fast_move_name="Vine Whip",
            charged_move_key="solar-beam",
            charged_move_name="Solar Beam",
        ),
        frame_index=609,
    )

    result = assemble_export_rows([[appraisal], [moves]])

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["species_key"] == "ivysaur"
    assert row["cp"] == 936
    assert row["hp_current"] == 50
    assert row["charged_move_name"] == "Solar Beam"


def test_assemble_export_rows_blanks_older_clean_cp_conflict() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(_detail_anchor_fields(cp=100), frame_index=1),
                _fragment(_detail_anchor_fields(cp=200), frame_index=20),
            ]
        ]
    )

    assert len(result.rows) == 1
    assert result.rows[0]["cp"] is None
    assert any(
        "Ignored conflicting optional export evidence for cp" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_uses_consensus_for_outlier_values() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(_core_fields(cp=689, iv_attack=15), frame_index=1),
                _fragment(_core_fields(cp=689, iv_attack=15), frame_index=2),
                _fragment(_core_fields(cp=689, iv_attack=15), frame_index=3),
                _fragment(_core_fields(cp=8866, iv_attack=14), frame_index=4),
            ]
        ]
    )

    assert not result.rows
    messages = [warning.message for warning in result.warnings]
    assert any("cp" in message for message in messages)
    assert any("iv_attack" in message for message in messages)


def test_assemble_export_rows_rejects_missing_weight_with_species_identity() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(
                    {
                        "species_key": "aggron",
                        "species_name": "Aggron",
                        "pokedex_id": 306,
                        "canonical_name_text": "Aggron",
                        "catch_date_text": "2/27/2023",
                        "location_text": "Prague, Czechia",
                        "cp": 1014,
                        "hp_current": 83,
                        "hp_max": 83,
                        "iv_attack": 14,
                        "iv_defense": 3,
                        "iv_stamina": 5,
                    },
                    frame_index=221,
                    fragment_type="appraisal",
                )
            ],
            [
                _fragment(
                    {
                        "species_key": "aggron",
                        "species_name": "Aggron",
                        "pokedex_id": 306,
                        "cp": 1014,
                        "hp_current": 83,
                        "hp_max": 83,
                        "fast_move_key": "smack-down-fast",
                        "fast_move_name": "Smack Down",
                        "charged_move_key": "frustration",
                        "charged_move_name": "Frustration",
                        "is_shadow": True,
                    },
                    frame_index=449,
                )
            ],
        ]
    )

    assert not result.rows
    assert result.rejected_sequence_count == 2
    assert any(
        "missing core identity field(s): weight_kg" in warning.message
        for warning in result.warnings
    )


def test_assemble_export_rows_prefers_gigantamax_over_dynamax() -> None:
    result = assemble_export_rows(
        [
            [
                _fragment(
                    _detail_anchor_fields(cp=1446, has_dynamax=True),
                    frame_index=1,
                ),
                _fragment({"has_gigantamax": True}, frame_index=2),
            ]
        ]
    )

    row = result.rows[0]
    assert row["has_gigantamax"] is True
    assert row["has_dynamax"] is False


def test_assemble_export_rows_rejects_critical_conflict_for_same_core_identity() -> (
    None
):
    result = assemble_export_rows(
        [
            [
                _fragment(
                    _detail_anchor_fields(cp=100, canonical_name_text="Bulbasaur"),
                    frame_index=1,
                )
            ],
            [
                _fragment(
                    _detail_anchor_fields(
                        cp=100,
                        canonical_name_text="Charmander",
                    ),
                    frame_index=20,
                )
            ],
        ]
    )

    assert not result.rows
    assert result.rejected_sequence_count == 2
    assert any(
        "Rejected conflicting export evidence for canonical_name" in warning.message
        for warning in result.warnings
    )


def test_export_writers_emit_csv_and_xlsx(tmp_path: Path) -> None:
    row = assemble_export_rows(
        [[_fragment(_detail_anchor_fields(cp=321, height_m=0.7, has_dynamax=True))]]
    ).rows[0]
    csv_path = tmp_path / "pokemon.csv"
    xlsx_path = tmp_path / "pokemon.xlsx"

    write_export_csv(csv_path, [row])
    write_export_xlsx(xlsx_path, [row])

    with csv_path.open(encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert csv_rows[0]["cp"] == "321"
    assert csv_rows[0]["height_m"] == "0.7"
    assert csv_rows[0]["has_dynamax"] == "True"
    assert list(csv_rows[0]) == list(EXPORT_COLUMNS)

    workbook = load_workbook(xlsx_path)
    worksheet = workbook["Pokemon"]
    assert worksheet["A1"].value == "source_file"
    assert worksheet["O2"].value == 321
    assert worksheet["S2"].value == 0.7
    assert worksheet["AJ2"].value is True


def _features(*enabled: str) -> dict[str, bool]:
    features = {key: False for key in FEATURE_KEYS}
    for key in enabled:
        features[key] = True
    return features


def _iv_evidence() -> _IvEvidence:
    return _empty_iv_evidence_from_signals({})


def _visual_record(frame_index: int) -> FrameVisualRecord:
    source = SourceAsset(Path("source.mp4"), "video")
    return FrameVisualRecord(
        frame=FrameCandidate(
            source, Path(f"frame_{frame_index:06d}.jpg"), frame_index, 0.0
        ),
        source_file="source.mp4",
        source_type="video",
        frame_path=f"frame_{frame_index:06d}.jpg",
        frame_index=frame_index,
        timestamp_s=float(frame_index),
        raw_classification="detail",
        signals={"stable_detail_signal": True},
        iv_evidence=_iv_evidence(),
        moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
        motion_sample=None,
    )


def _sequence_result(sequence: list[FrameVisualRecord]) -> ProductionSequenceScanResult:
    visual = sequence[0]
    record = FrameScanRecord(
        source_file=visual.source_file,
        source_type=visual.source_type,
        frame_path=visual.frame_path,
        frame_index=visual.frame_index,
        timestamp_s=visual.timestamp_s,
        classification="detail",
        raw_classification="detail",
        features=_features("has_CP", "has_hp", "has_weight"),
        values={
            "cp": visual.frame_index,
            "hp": "10/10",
            "weight_kg": f"{visual.frame_index / 100:.2f}",
        },
        ocr={},
    )
    return ProductionSequenceScanResult(
        records=[record],
        accepted_fields={
            "cp": visual.frame_index,
            "hp": "10/10",
            "weight": record.values["weight_kg"],
        },
        desired_fields={"cp", "hp", "weight"},
        requested_ocr_fields_by_frame={visual.frame_index: ()},
        completed=True,
    )


def _near_miss_sequence_result(
    sequence: list[FrameVisualRecord],
) -> ProductionSequenceScanResult:
    base_result = _sequence_result(sequence)
    record = base_result.records[0]
    return ProductionSequenceScanResult(
        records=[record],
        accepted_fields={"hp": "10/10", "weight": record.values["weight_kg"]},
        desired_fields={"hp", "weight", "moves"},
        requested_ocr_fields_by_frame={},
        completed=False,
        sequence_type="detail/raw=appraisal",
    )


def _bounded_video_frames(
    source_asset: SourceAsset,
    source_frames_dir: Path,
    frame_indexes: range,
    *,
    total_frame_count: int,
    duration_s: float,
) -> export_module.VideoExtractionResult:
    source_frames_dir.mkdir(parents=True, exist_ok=True)
    frames: list[FrameCandidate] = []
    denominator = max(1, total_frame_count - 1)
    for frame_index in frame_indexes:
        frame_path = export_module._video_frame_path(source_frames_dir, frame_index)
        frame_path.write_text(f"frame {frame_index}", encoding="utf-8")
        timestamp_s = duration_s * frame_index / denominator if duration_s else 0.0
        frames.append(
            FrameCandidate(source_asset, frame_path, frame_index, timestamp_s)
        )
    return export_module.VideoExtractionResult(frames, "test", [])


def _row_compatible_scan_result(
    sequence: list[FrameVisualRecord],
) -> ProductionSequenceScanResult:
    records: list[FrameScanRecord] = []
    for visual in sorted(sequence, key=lambda record: record.frame_index):
        records.append(
            FrameScanRecord(
                source_file=visual.source_file,
                source_type=visual.source_type,
                frame_path=visual.frame_path,
                frame_index=visual.frame_index,
                timestamp_s=visual.timestamp_s,
                classification="detail",
                raw_classification="detail",
                features=_features("has_CP", "has_hp", "has_weight", "has_moves"),
                values={"cp": 123, "hp": "10/10", "weight_kg": "1.25"},
                ocr={"moves": {"text": "Vine Whip Solar Beam", "confidence": 0.90}},
            )
        )
    return ProductionSequenceScanResult(
        records=records,
        accepted_fields={"cp": 123, "hp": "10/10", "weight": "1.25"},
        desired_fields={"cp", "hp", "weight", "moves"},
        requested_ocr_fields_by_frame={record.frame_index: () for record in records},
        completed=True,
        sequence_type="detail/raw=detail",
    )


def test_stabilize_same_hp_sequence_weights_refreshes_fields() -> None:
    source = SourceAsset(Path("source.mp4"), "video")

    def record(index: int, weight: str, *, moves: str = "") -> FrameScanRecord:
        frame = FrameCandidate(source, Path(f"frame_{index:06d}.png"), index, 0.0)
        features = _features("has_hp", "has_weight")
        if moves:
            features["has_moves"] = True
        return FrameScanRecord(
            source_file=source.path.name,
            source_type=source.source_type,
            frame_path=str(frame.frame_path),
            frame_index=index,
            timestamp_s=frame.timestamp_s,
            classification="detail",
            raw_classification="detail",
            features=features,
            values={"hp": "66/66", "weight_kg": weight},
            ocr={"moves": {"text": moves, "confidence": 0.90}},
        )

    def result(records: list[FrameScanRecord]) -> ProductionSequenceScanResult:
        return ProductionSequenceScanResult(
            records=records,
            accepted_fields={"hp": "66/66", "weight": records[0].values["weight_kg"]},
            desired_fields={"hp", "weight", "moves"},
            requested_ocr_fields_by_frame={},
            completed=False,
            sequence_type="detail/raw=detail",
        )

    move_record = record(372, "91.87", moves="Air Slash Frustration")
    move_result = result([move_record])
    physical_result = result(
        [record(380, "51.87"), record(381, "51.87"), record(382, "51.87")]
    )

    warnings = export_module._stabilize_same_hp_sequence_weights(
        [move_result, physical_result]
    )

    assert move_record.values["weight_kg"] == "51.87"
    assert move_record.features["has_weight"]
    assert move_record.signals["cross_sequence_weight_corrected"] is True
    assert move_result.accepted_fields["weight"] == "51.87"
    assert move_result.accepted_fields["moves"] == "Air Slash Frustration"
    assert warnings
    assert warnings[0].kind == "sequence_weight"
    assert warnings[0].first_frame_index == 372
    assert warnings[0].last_frame_index == 382


def test_sequence_scanning_parallel_path_preserves_sequence_order(
    tmp_path: Path,
) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    second_sequence_started = threading.Event()

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        if sequence[0].frame_index == 20:
            second_sequence_started.set()
        else:
            assert second_sequence_started.wait(timeout=5)
        return _sequence_result(sequence)

    result = _scan_production_sequences(
        sequences,
        ScanSettings(Path("."), tmp_path, workers=2),
        sequence_scanner,
    )

    assert result.worker_count == 2
    assert [scan.records[0].frame_index for scan in result.records] == [10, 20]
    assert result.retry_count == 0
    assert not result.warnings


def test_sequence_scanning_retries_failed_sequence(tmp_path: Path) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    attempts: dict[int, int] = {}

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        frame_index = sequence[0].frame_index
        attempts[frame_index] = attempts.get(frame_index, 0) + 1
        if frame_index == 10 and attempts[frame_index] == 1:
            msg = "temporary sequence failure"
            raise RuntimeError(msg)
        return _sequence_result(sequence)

    result = _scan_production_sequences(
        sequences,
        ScanSettings(Path("."), tmp_path, workers=2, max_frame_attempts=2),
        sequence_scanner,
    )

    assert [scan.records[0].frame_index for scan in result.records] == [10, 20]
    assert result.worker_count == 2
    assert result.retry_count == 1
    assert attempts == {10: 2, 20: 1}
    assert any(
        "requeued with reduced concurrency" in warning.message
        for warning in result.warnings
    )
    assert not any(
        "Skipped production sequence" in warning.message for warning in result.warnings
    )


def test_sequence_scanning_skips_exhausted_failure(tmp_path: Path) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        if sequence[0].frame_index == 10:
            msg = "permanent sequence failure"
            raise RuntimeError(msg)
        return _sequence_result(sequence)

    result = _scan_production_sequences(
        sequences,
        ScanSettings(Path("."), tmp_path, workers=2, max_frame_attempts=2),
        sequence_scanner,
    )

    assert [scan.records[0].frame_index for scan in result.records] == [20]
    assert result.retry_count == 1
    assert any(
        "requeued with reduced concurrency" in warning.message
        for warning in result.warnings
    )
    failure_warnings = [
        warning
        for warning in result.warnings
        if "Skipped production sequence" in warning.message
    ]
    assert len(failure_warnings) == 1
    assert failure_warnings[0].source_file == "source.mp4"
    assert failure_warnings[0].first_frame_index == 10
    assert failure_warnings[0].last_frame_index == 10


def test_repair_production_sequences_repairs_only_near_miss_identity(
    tmp_path: Path,
    monkeypatch,
) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    near_miss = _near_miss_sequence_result(sequences[0])
    complete = _sequence_result(sequences[1])
    repaired_result = _sequence_result(sequences[0])
    repaired_sequences: list[list[FrameVisualRecord]] = []

    def repair_scanner(
        repair_sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        repaired_sequences.append(repair_sequence)
        return repaired_result

    monkeypatch.setattr(
        export_module,
        "scan_production_sequence_repair",
        repair_scanner,
    )

    result = export_module._repair_production_sequences(
        sequences,
        [near_miss, complete],
        ScanSettings(Path("."), tmp_path),
    )

    assert result.records == [repaired_result, complete]
    assert result.repaired_count == 1
    assert result.worker_count == 1
    assert result.retry_count == 0
    assert repaired_sequences == [sequences[0]]


def test_repair_production_sequences_parallel_path_preserves_order(
    tmp_path: Path,
    monkeypatch,
) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    second_sequence_started = threading.Event()

    def repair_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        if sequence[0].frame_index == 20:
            second_sequence_started.set()
        else:
            assert second_sequence_started.wait(timeout=5)
        return _sequence_result(sequence)

    monkeypatch.setattr(
        export_module,
        "scan_production_sequence_repair",
        repair_scanner,
    )

    result = export_module._repair_production_sequences(
        sequences,
        [_near_miss_sequence_result(sequence) for sequence in sequences],
        ScanSettings(Path("."), tmp_path, workers=2),
    )

    assert result.worker_count == 2
    assert [scan.records[0].frame_index for scan in result.records] == [10, 20]
    assert result.repaired_count == 2
    assert result.retry_count == 0
    assert not result.warnings


def test_repair_production_sequences_retries_failed_repair(
    tmp_path: Path,
    monkeypatch,
) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    attempts: dict[int, int] = {}

    def repair_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        frame_index = sequence[0].frame_index
        attempts[frame_index] = attempts.get(frame_index, 0) + 1
        if frame_index == 10 and attempts[frame_index] == 1:
            msg = "temporary repair failure"
            raise RuntimeError(msg)
        return _sequence_result(sequence)

    monkeypatch.setattr(
        export_module,
        "scan_production_sequence_repair",
        repair_scanner,
    )

    result = export_module._repair_production_sequences(
        sequences,
        [_near_miss_sequence_result(sequence) for sequence in sequences],
        ScanSettings(Path("."), tmp_path, workers=2, max_frame_attempts=2),
    )

    assert [scan.records[0].frame_index for scan in result.records] == [10, 20]
    assert result.worker_count == 2
    assert result.retry_count == 1
    assert result.repaired_count == 2
    assert attempts == {10: 2, 20: 1}
    assert any(
        warning.kind == "sequence_repair"
        and "requeued with reduced concurrency" in warning.message
        for warning in result.warnings
    )


def test_repair_production_sequences_keeps_original_after_exhausted_failure(
    tmp_path: Path,
    monkeypatch,
) -> None:
    sequences = [[_visual_record(10)], [_visual_record(20)]]
    initial_results = [_near_miss_sequence_result(sequence) for sequence in sequences]

    def repair_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        if sequence[0].frame_index == 10:
            msg = "permanent repair failure"
            raise RuntimeError(msg)
        return _sequence_result(sequence)

    monkeypatch.setattr(
        export_module,
        "scan_production_sequence_repair",
        repair_scanner,
    )

    result = export_module._repair_production_sequences(
        sequences,
        initial_results,
        ScanSettings(Path("."), tmp_path, workers=2, max_frame_attempts=2),
    )

    assert result.records[0] is initial_results[0]
    assert result.records[1].records[0].frame_index == 20
    assert result.repaired_count == 1
    assert result.retry_count == 1
    failure_warnings = [
        warning for warning in result.warnings if "repair failure" in warning.message
    ]
    assert len(failure_warnings) == 1
    assert failure_warnings[0].kind == "sequence_repair"
    assert failure_warnings[0].source_file == "source.mp4"
    assert failure_warnings[0].first_frame_index == 10


def test_run_production_export_with_injected_scanners(tmp_path: Path) -> None:
    input_path = tmp_path / "frame.png"
    output_dir = tmp_path / "export"
    Image.new("RGB", (8, 8), "white").save(input_path)

    def visual_scanner(frame: FrameCandidate) -> FrameVisualRecord:
        return FrameVisualRecord(
            frame=frame,
            source_file=frame.source_asset.path.name,
            source_type=frame.source_asset.source_type,
            frame_path=str(frame.frame_path),
            frame_index=frame.frame_index,
            timestamp_s=frame.timestamp_s,
            raw_classification="detail",
            signals={"stable_detail_signal": True},
            iv_evidence=_iv_evidence(),
            moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
            motion_sample=None,
        )

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        visual = sequence[0]
        record = FrameScanRecord(
            source_file=visual.source_file,
            source_type=visual.source_type,
            frame_path=visual.frame_path,
            frame_index=visual.frame_index,
            timestamp_s=visual.timestamp_s,
            classification="detail",
            raw_classification="detail",
            features=_features("has_CP", "has_hp", "has_weight", "has_moves"),
            values={"cp": 123, "hp": "10/10", "weight_kg": "1.25"},
            ocr={"moves": {"text": "Vine Whip Solar Beam", "confidence": 0.90}},
        )
        return ProductionSequenceScanResult(
            records=[record],
            accepted_fields={"cp": 123, "hp": "10/10", "weight": "1.25"},
            desired_fields={"cp", "hp", "weight"},
            requested_ocr_fields_by_frame={visual.frame_index: ()},
            completed=True,
        )

    report = run_production_export(
        ScanSettings(input_path=input_path, output_dir=output_dir, workers=1),
        visual_scanner=visual_scanner,
        sequence_scanner=sequence_scanner,
    )

    assert len(report.rows) == 1
    assert report.rows[0]["cp"] == 123
    assert (output_dir / "pokemon.csv").exists()
    assert (output_dir / "pokemon.xlsx").exists()
    live_log = output_dir / "export.log"
    assert live_log.exists()
    live_log_lines = live_log.read_text(encoding="utf-8").splitlines()
    assert (
        live_log_lines[0]
        == "timestamp\tworker_id\tphase\tframe_name\tsource_name\tfields"
    )
    assert live_log_lines[1].split("\t")[2:4] == ["visual", "frame_000000.png"]

    manifest = json.loads(
        (output_dir / "artifacts" / "export_manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["exported_row_count"] == 1
    assert manifest["sequence_worker_count"] == 1
    assert manifest["sequence_retry_count"] == 0
    assert manifest["repair_worker_count"] == 1
    assert manifest["repair_retry_count"] == 0
    assert manifest["repaired_sequence_count"] == 0
    assert manifest["timing_summary"]["sequence_worker_count"] == 1
    assert manifest["timing_summary"]["sequence_retry_count"] == 0
    assert manifest["timing_summary"]["repair_worker_count"] == 1
    assert manifest["timing_summary"]["repair_retry_count"] == 0
    assert manifest["timing_summary"]["repaired_sequence_count"] == 0
    assert manifest["artifacts"]["pokemon_csv"] == str(output_dir / "pokemon.csv")
    assert manifest["artifacts"]["export_log"] == str(live_log)


def test_bounded_export_delete_helper_refuses_original_and_outside_files(
    tmp_path: Path,
) -> None:
    source_file = tmp_path / "original.png"
    source_file.write_text("source", encoding="utf-8")
    source_frames_dir = tmp_path / "artifacts" / "clip" / "frames"
    source_frames_dir.mkdir(parents=True)

    original_frame = FrameCandidate(
        SourceAsset(source_file, "image"),
        source_file,
        0,
        0.0,
    )
    assert (
        export_module._delete_export_frame_file(original_frame, source_frames_dir)
        is False
    )
    assert source_file.exists()

    outside_frame_path = tmp_path / "outside" / "frame_000001.png"
    outside_frame_path.parent.mkdir()
    outside_frame_path.write_text("outside", encoding="utf-8")
    outside_frame = FrameCandidate(
        SourceAsset(tmp_path / "clip.mp4", "video"),
        outside_frame_path,
        0,
        0.0,
    )
    assert (
        export_module._delete_export_frame_file(outside_frame, source_frames_dir)
        is False
    )
    assert outside_frame_path.exists()

    inside_frame_path = source_frames_dir / "frame_000001.png"
    inside_frame_path.write_text("inside", encoding="utf-8")
    inside_frame = FrameCandidate(
        SourceAsset(tmp_path / "clip.mp4", "video"),
        inside_frame_path,
        0,
        0.0,
    )
    assert (
        export_module._delete_export_frame_file(inside_frame, source_frames_dir) is True
    )
    assert not inside_frame_path.exists()


def test_bounded_cleanup_deletes_unsequenced_visual_frame_with_reason(
    tmp_path: Path,
) -> None:
    source_frames_dir = tmp_path / "artifacts" / "clip" / "frames"
    source_frames_dir.mkdir(parents=True)
    source = SourceAsset(tmp_path / "clip.mp4", "video")
    records: list[FrameVisualRecord] = []
    for frame_index in range(2):
        frame_path = source_frames_dir / f"frame_{frame_index:06d}.png"
        frame_path.write_text("frame", encoding="utf-8")
        frame = FrameCandidate(source, frame_path, frame_index, float(frame_index))
        records.append(
            FrameVisualRecord(
                frame=frame,
                source_file=source.path.name,
                source_type=source.source_type,
                frame_path=str(frame_path),
                frame_index=frame_index,
                timestamp_s=float(frame_index),
                raw_classification="detail",
                signals={"stable_detail_signal": True},
                iv_evidence=_iv_evidence(),
                moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
                motion_sample=None,
            )
        )
    live_log = export_module._ExportLiveLogger(tmp_path / "export.log")
    stats = export_module._BoundedFrameFileStats(max_export_frame_files=1)

    export_module._delete_unsequenced_visual_frame_files(
        records,
        {0},
        source_frames_dir,
        stats,
        live_log,
    )

    assert records[0].frame.frame_path.exists()
    assert not records[1].frame.frame_path.exists()
    assert stats.deleted_unsequenced_visual_frames == 1
    assert any(
        event.action == "deleted"
        and event.reason == "visual_not_in_completed_sequence"
        and event.frame_index == 1
        for event in stats.frame_lifecycle_events
    )


def test_bounded_sequence_files_survive_repair_then_release(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_frames_dir = tmp_path / "artifacts" / "clip" / "frames"
    source_frames_dir.mkdir(parents=True)
    frame_path = source_frames_dir / "frame_000001.png"
    frame_path.write_text("frame", encoding="utf-8")
    source = SourceAsset(tmp_path / "clip.mp4", "video")
    frame = FrameCandidate(source, frame_path, 0, 0.0)
    sequence = [
        FrameVisualRecord(
            frame=frame,
            source_file=source.path.name,
            source_type=source.source_type,
            frame_path=str(frame_path),
            frame_index=0,
            timestamp_s=0.0,
            raw_classification="detail",
            signals={"stable_detail_signal": True},
            iv_evidence=_iv_evidence(),
            moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
            motion_sample=None,
        )
    ]
    repair_saw_file = False

    def initial_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        assert sequence[0].frame.frame_path.exists()
        result = _row_compatible_scan_result(sequence)
        result.accepted_fields = {"hp": "10/10", "weight": "1.25"}
        result.completed = False
        return result

    def repair_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
        *,
        progress_callback=None,
    ) -> ProductionSequenceScanResult:
        del settings, progress_callback
        nonlocal repair_saw_file
        repair_saw_file = sequence[0].frame.frame_path.exists()
        return _row_compatible_scan_result(sequence)

    monkeypatch.setattr(export_module, "scan_production_sequence", initial_scanner)
    monkeypatch.setattr(
        export_module, "scan_production_sequence_repair", repair_scanner
    )

    stats = export_module._BoundedFrameFileStats(max_export_frame_files=1)

    def release_sequence(sequence: list[FrameVisualRecord]) -> None:
        export_module._delete_sequence_frame_files(sequence, source_frames_dir, stats)

    lifecycle = export_module._scan_sequences_with_optional_repair(
        [(0, 0, sequence)],
        ScanSettings(source.path, tmp_path / "out", workers=1),
        export_module.scan_production_sequence,
        release_callback=release_sequence,
    )

    assert repair_saw_file is True
    assert lifecycle.repaired_count == 1
    assert stats.deleted_sequence_frames == 1
    assert not frame_path.exists()


def test_video_frame_timeline_probe_accepts_vfr_timestamps(
    tmp_path: Path, monkeypatch
) -> None:
    def fake_run(command, *, check, capture_output, text):
        del command, check, capture_output, text

        class Completed:
            stdout = json.dumps(
                {
                    "frames": [
                        {"best_effort_timestamp_time": "0.000000"},
                        {"best_effort_timestamp_time": "0.040000"},
                        {"best_effort_timestamp_time": "0.120000"},
                        {"best_effort_timestamp_time": "0.160000"},
                    ]
                }
            )

        return Completed()

    monkeypatch.setattr(export_module.subprocess, "run", fake_run)

    probe = export_module._probe_video_frame_timeline(tmp_path / "clip.mp4", 4)

    assert probe.status == "available"
    assert probe.frame_count == 4
    assert probe.timeline is not None
    assert probe.timeline.timestamps_s == (0.0, 0.04, 0.12, 0.16)


def test_video_frame_timeline_probe_rejects_incomplete_map(
    tmp_path: Path, monkeypatch
) -> None:
    def fake_run(command, *, check, capture_output, text):
        del command, check, capture_output, text

        class Completed:
            stdout = json.dumps(
                {
                    "frames": [
                        {"best_effort_timestamp_time": "0.000000"},
                        {"best_effort_timestamp_time": "0.040000"},
                    ]
                }
            )

        return Completed()

    monkeypatch.setattr(export_module.subprocess, "run", fake_run)

    probe = export_module._probe_video_frame_timeline(tmp_path / "clip.mp4", 3)

    assert probe.status == "unavailable"
    assert probe.frame_count == 2
    assert probe.reason == "frame_count_mismatch:expected=3:actual=2"
    assert probe.timeline is None


def test_time_guided_chunk_extracts_to_temp_then_canonical_names(
    tmp_path: Path, monkeypatch
) -> None:
    source = SourceAsset(tmp_path / "clip.mp4", "video")
    source_frames_dir = tmp_path / "frames"
    commands: list[list[str]] = []

    def fake_run(command, *, check, capture_output, text):
        del check, capture_output, text
        commands.append(command)
        start_number = int(command[command.index("-start_number") + 1])
        pattern = Path(command[-1])
        pattern.parent.mkdir(parents=True, exist_ok=True)
        for number in range(start_number, start_number + 3):
            (pattern.parent / f"frame_{number:06d}.png").write_text(
                f"frame {number}", encoding="utf-8"
            )

    monkeypatch.setattr(export_module.subprocess, "run", fake_run)

    result = export_module._extract_video_frame_chunk(
        source,
        source_frames_dir,
        2,
        4,
        total_frame_count=6,
        duration_s=5.0,
        timeline=export_module._VideoFrameTimeline((0.0, 1.0, 2.0, 3.0, 4.0, 5.0)),
    )

    assert result.method == "time_seek"
    assert result.fallback_reason == ""
    assert result.seek_start_s is not None
    assert result.seek_duration_s is not None
    assert [frame.frame_index for frame in result.extraction.frames] == [2, 3, 4]
    assert [path.name for path in sorted(source_frames_dir.glob("frame_*.png"))] == [
        "frame_000003.png",
        "frame_000004.png",
        "frame_000005.png",
    ]
    assert not list(source_frames_dir.glob("__bounded_chunk_*"))
    assert "-ss" in commands[0]
    assert "__bounded_chunk_2_4" in str(commands[0][-1])


def test_time_guided_chunk_falls_back_when_temp_output_mismatches(
    tmp_path: Path, monkeypatch
) -> None:
    source = SourceAsset(tmp_path / "clip.mp4", "video")
    source_frames_dir = tmp_path / "frames"
    command_methods: list[str] = []

    def fake_run(command, *, check, capture_output, text):
        del check, capture_output, text
        start_number = int(command[command.index("-start_number") + 1])
        pattern = Path(command[-1])
        pattern.parent.mkdir(parents=True, exist_ok=True)
        if "-ss" in command:
            command_methods.append("time_seek")
            count = 2
        else:
            command_methods.append("range_select")
            count = 3
        for number in range(start_number, start_number + count):
            (pattern.parent / f"frame_{number:06d}.png").write_text(
                f"frame {number}", encoding="utf-8"
            )

    monkeypatch.setattr(export_module.subprocess, "run", fake_run)

    result = export_module._extract_video_frame_chunk(
        source,
        source_frames_dir,
        2,
        4,
        total_frame_count=6,
        duration_s=5.0,
        timeline=export_module._VideoFrameTimeline((0.0, 1.0, 2.0, 3.0, 4.0, 5.0)),
    )

    assert result.method == "range_select"
    assert result.fallback_reason == "time_seek_validation_failed"
    assert command_methods == ["time_seek", "range_select"]
    assert [frame.frame_index for frame in result.extraction.frames] == [2, 3, 4]
    assert [path.name for path in sorted(source_frames_dir.glob("frame_*.png"))] == [
        "frame_000003.png",
        "frame_000004.png",
        "frame_000005.png",
    ]


def test_time_guided_chunks_match_range_select_on_generated_video(
    tmp_path: Path,
) -> None:
    if shutil.which("ffmpeg") is None or shutil.which("ffprobe") is None:
        pytest.skip("FFmpeg/FFprobe are required for chunk equivalence regression.")

    frame_source_dir = tmp_path / "source_frames"
    frame_source_dir.mkdir()
    for frame_index in range(12):
        image = Image.new(
            "RGB",
            (96, 64),
            (
                frame_index * 20 % 255,
                frame_index * 40 % 255,
                frame_index * 60 % 255,
            ),
        )
        image.save(frame_source_dir / f"source_{frame_index + 1:03d}.png")

    video_path = tmp_path / "clip.mp4"
    export_module.subprocess.run(
        [
            "ffmpeg",
            "-y",
            "-framerate",
            "4",
            "-i",
            str(frame_source_dir / "source_%03d.png"),
            "-c:v",
            "mpeg4",
            "-q:v",
            "2",
            str(video_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    timeline_probe = export_module._probe_video_frame_timeline(video_path, 12)
    assert timeline_probe.timeline is not None
    source = SourceAsset(video_path, "video")

    for first_frame_index, last_frame_index in ((0, 3), (4, 7), (8, 11)):
        exact = export_module._extract_video_frame_chunk(
            source,
            tmp_path / f"exact_{first_frame_index}",
            first_frame_index,
            last_frame_index,
            total_frame_count=12,
            duration_s=3.0,
        )
        fast = export_module._extract_video_frame_chunk(
            source,
            tmp_path / f"fast_{first_frame_index}",
            first_frame_index,
            last_frame_index,
            total_frame_count=12,
            duration_s=3.0,
            timeline=timeline_probe.timeline,
        )

        assert fast.method == "time_seek"
        assert [frame.frame_index for frame in fast.extraction.frames] == list(
            range(first_frame_index, last_frame_index + 1)
        )
        for exact_frame, fast_frame in zip(
            exact.extraction.frames, fast.extraction.frames, strict=True
        ):
            assert (
                exact_frame.frame_path.read_bytes()
                == fast_frame.frame_path.read_bytes()
            )


# pylint: disable-next=too-many-statements
def test_bounded_export_matches_unlimited_and_releases_frame_files(
    tmp_path: Path,
    monkeypatch,
) -> None:
    input_path = tmp_path / "clip.mp4"
    input_path.write_text("fake video", encoding="utf-8")
    total_frame_count = 8
    duration_s = 7.0
    sequence_calls: list[list[int]] = []

    def fake_extract_video_frames(
        source_asset: SourceAsset,
        source_frames_dir: Path,
    ) -> export_module.VideoExtractionResult:
        return _bounded_video_frames(
            source_asset,
            source_frames_dir,
            range(total_frame_count),
            total_frame_count=total_frame_count,
            duration_s=duration_s,
        )

    def fake_extract_video_frame_chunk(
        source_asset: SourceAsset,
        source_frames_dir: Path,
        first_frame_index: int,
        last_frame_index: int,
        *,
        total_frame_count: int,
        duration_s: float,
        timeline=None,
    ) -> export_module._BoundedChunkExtractionResult:
        del timeline
        extraction = _bounded_video_frames(
            source_asset,
            source_frames_dir,
            range(first_frame_index, last_frame_index + 1),
            total_frame_count=total_frame_count,
            duration_s=duration_s,
        )
        return export_module._BoundedChunkExtractionResult(
            extraction,
            "time_seek",
            seek_start_s=float(first_frame_index),
            seek_duration_s=float(last_frame_index - first_frame_index + 1),
        )

    def visual_scanner(frame: FrameCandidate) -> FrameVisualRecord:
        raw_classification = "detail"
        stable = True
        if frame.frame_index == 7:
            raw_classification = "list"
            stable = False
        elif frame.frame_index == 6:
            raw_classification = NON_EXTRACTABLE_CLASS
            stable = False
        return FrameVisualRecord(
            frame=frame,
            source_file=frame.source_asset.path.name,
            source_type=frame.source_asset.source_type,
            frame_path=str(frame.frame_path),
            frame_index=frame.frame_index,
            timestamp_s=frame.timestamp_s,
            raw_classification=raw_classification,
            signals={"stable_detail_signal": stable},
            iv_evidence=_iv_evidence(),
            moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
            motion_sample=None,
        )

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        frame_indexes = sorted(record.frame_index for record in sequence)
        assert frame_indexes == list(range(6))
        assert all(record.frame.frame_path.exists() for record in sequence)
        sequence_calls.append(frame_indexes)
        return _row_compatible_scan_result(sequence)

    monkeypatch.setattr(
        export_module, "extract_video_frames", fake_extract_video_frames
    )
    monkeypatch.setattr(export_module, "_probe_video_frame_count", lambda _path: 8)
    monkeypatch.setattr(export_module, "probe_video_duration", lambda _path: duration_s)
    monkeypatch.setattr(
        export_module,
        "_probe_video_frame_timeline",
        lambda _path, _count: export_module._VideoFrameTimelineProbe(
            "available",
            frame_count=total_frame_count,
            timeline=export_module._VideoFrameTimeline(
                tuple(float(index) for index in range(total_frame_count))
            ),
        ),
    )
    monkeypatch.setattr(
        export_module,
        "_extract_video_frame_chunk",
        fake_extract_video_frame_chunk,
    )

    unlimited_report = run_production_export(
        ScanSettings(
            input_path=input_path,
            output_dir=tmp_path / "unlimited",
            workers=1,
        ),
        visual_scanner=visual_scanner,
        sequence_scanner=sequence_scanner,
    )
    bounded_report = run_production_export(
        ScanSettings(
            input_path=input_path,
            output_dir=tmp_path / "bounded",
            workers=1,
            max_export_frame_files=3,
        ),
        visual_scanner=visual_scanner,
        sequence_scanner=sequence_scanner,
    )

    assert sequence_calls == [list(range(6)), list(range(6))]
    assert bounded_report.rows == unlimited_report.rows
    assert input_path.exists()
    assert bounded_report.bounded_extraction_enabled is True
    assert bounded_report.max_export_frame_files == 3
    assert bounded_report.peak_export_frame_files > 3
    assert bounded_report.deleted_list_or_non_extractable_frames == 2
    assert bounded_report.deleted_sequence_frames == 6
    assert bounded_report.bounded_extraction_soft_limit_exceeded is True

    frames_dir = tmp_path / "bounded" / "artifacts" / "clip" / "frames"
    assert not list(frames_dir.glob("frame_*.png"))

    manifest = json.loads(
        (tmp_path / "bounded" / "artifacts" / "export_manifest.json").read_text(
            encoding="utf-8"
        )
    )
    assert manifest["max_export_frame_files"] == 3
    assert manifest["bounded_extraction_enabled"] is True
    assert manifest["peak_export_frame_files"] > 3
    assert manifest["deleted_list_or_non_extractable_frames"] == 2
    assert manifest["deleted_sequence_frames"] == 6
    assert manifest["deleted_unsequenced_visual_frames"] == 0
    assert manifest["retained_frame_count"] == 0
    assert manifest["frame_lifecycle_summary"]["actions"]["deleted"] == 8
    assert manifest["bounded_extraction_soft_limit_exceeded"] is True
    assert manifest["timing_summary"]["bounded_extraction_enabled"] is True
    assert manifest["artifacts"]["frame_lifecycle_jsonl"] == str(
        tmp_path / "bounded" / "artifacts" / "frame_lifecycle.jsonl"
    )
    assert manifest["artifacts"]["performance_summary"] == str(
        tmp_path / "bounded" / "artifacts" / "performance_summary.json"
    )
    performance_summary = json.loads(
        (tmp_path / "bounded" / "artifacts" / "performance_summary.json").read_text(
            encoding="utf-8"
        )
    )
    assert performance_summary["bounded_frame_files"]["max_export_frame_files"] == 3
    assert performance_summary["bounded_frame_files"]["peak_export_frame_files"] > 3
    assert performance_summary["bounded_frame_files"]["chunk_count"] >= 1
    assert performance_summary["bounded_frame_files"]["seeked_chunk_count"] >= 1
    assert performance_summary["bounded_frame_files"]["fallback_chunk_count"] == 0
    assert all(
        chunk["extraction_method"] == "time_seek"
        for chunk in performance_summary["bounded_frame_files"]["chunks"]
    )
    assert (
        performance_summary["bounded_frame_files"]["chunks"][-1][
            "frame_files_after_final_cleanup"
        ]
        == 0
    )
    assert performance_summary["workers"]["configured"] == 1
    assert (
        performance_summary["workers"]["summary_by_phase"]["visual_analysis"][
            "batch_count"
        ]
        >= 1
    )
    assert performance_summary["frames"]["deleted"] == 8
    video_extraction = manifest["sources"]["clip.mp4"]["video_extraction"]
    assert video_extraction["timeline_probe"]["status"] == "available"
    assert video_extraction["seeked_chunk_count"] >= 1
    assert video_extraction["fallback_chunk_count"] == 0
    export_log = (tmp_path / "bounded" / "export.log").read_text(encoding="utf-8")
    assert "worker active_worker_count=1" in export_log
    assert "bounded chunk accounting" in export_log
    assert "method=time_seek" in export_log
    lifecycle_rows = [
        json.loads(line)
        for line in (tmp_path / "bounded" / "artifacts" / "frame_lifecycle.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if line
    ]
    assert any(
        row["action"] == "deleted" and row["reason"] == "completed_sequence_released"
        for row in lifecycle_rows
    )

    warning_rows = [
        json.loads(line)
        for line in (tmp_path / "bounded" / "artifacts" / "warnings.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if line
    ]
    assert any(warning["kind"] == "bounded_extraction" for warning in warning_rows)


def test_bounded_iaast_export_matches_unlimited_toxtricity_iv(tmp_path: Path) -> None:
    input_path = Path("example") / "screen-20260426-182559-1777220738734_iaast.mp4"
    if not input_path.exists():
        pytest.skip("IAast example video is not available.")
    if shutil.which("ffmpeg") is None or shutil.which("ffprobe") is None:
        pytest.skip("FFmpeg/FFprobe are required for the IAast export regression.")
    if not export_module.TesseractOcrEngine(lang="eng").is_available():
        pytest.skip("Tesseract is required for the IAast export regression.")

    unlimited_report = run_production_export(
        ScanSettings(
            input_path=input_path,
            output_dir=tmp_path / "unlimited",
        )
    )
    bounded_report = run_production_export(
        ScanSettings(
            input_path=input_path,
            output_dir=tmp_path / "bounded",
            max_export_frame_files=400,
        )
    )

    assert len(unlimited_report.rows) == 5
    assert len(bounded_report.rows) == 5
    unlimited_toxtricity = _row_by_species(unlimited_report.rows, "toxtricity")
    bounded_toxtricity = _row_by_species(bounded_report.rows, "toxtricity")
    for column in (*IV_NUMERIC_FIELD_NAMES, "appraisal_perfect"):
        assert bounded_toxtricity[column] == unlimited_toxtricity[column]

    manifest = json.loads(
        (tmp_path / "bounded" / "artifacts" / "export_manifest.json").read_text(
            encoding="utf-8"
        )
    )
    assert manifest["bounded_extraction_enabled"] is True
    assert manifest["max_export_frame_files"] == 400
    assert manifest["deleted_sequence_frames"] > 0
    assert bounded_toxtricity["iv_complete"] is True


def test_slow_litwick_bounded_export_matches_manual_reference(tmp_path: Path) -> None:
    if os.environ.get("POGO_SLOW_LITWICK_REGRESSION") != "1":
        pytest.skip("set POGO_SLOW_LITWICK_REGRESSION=1 to run the slow Litwick export")
    input_path = Path("example") / "screen-20260424-220745-1777061192595_litwick.mp4"
    manual_path = Path("output") / "2605312216_litwick_export" / "pokemonManual.xlsx"
    if not input_path.exists() or not manual_path.exists():
        pytest.skip("Litwick video and manual reference are required.")
    if shutil.which("ffmpeg") is None or shutil.which("ffprobe") is None:
        pytest.skip("FFmpeg/FFprobe are required for the Litwick export regression.")
    if not export_module.TesseractOcrEngine(lang="eng").is_available():
        pytest.skip("Tesseract is required for the Litwick export regression.")

    report = run_production_export(
        ScanSettings(
            input_path=input_path,
            output_dir=tmp_path / "litwick",
            max_export_frame_files=400,
        )
    )
    manual_rows = _xlsx_rows(manual_path)
    exported_keys = {_reference_identity_key(row) for row in report.rows}
    diagnostics = [
        json.loads(line)
        for line in (tmp_path / "litwick" / "artifacts" / "row_diagnostics.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if line
    ]
    unresolved = [
        row
        for row in diagnostics
        if row["outcome"] == "rejected" and row["pokemon_like"]
    ]

    missing = [
        row for row in manual_rows if _reference_identity_key(row) not in exported_keys
    ]
    assert not missing or unresolved
    assert all(row["iv_complete"] is True for row in report.rows if row["iv_sum"])
    assert not list((tmp_path / "litwick" / "artifacts").glob("**/frames/frame_*.png"))


def _xlsx_rows(path: Path) -> list[dict[str, export_module.ExportValue]]:
    worksheet = load_workbook(path, data_only=True).active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    header = [str(value) for value in rows[0]]
    return [
        {
            column: _export_cell_value(value)
            for column, value in zip(header, row, strict=True)
        }
        for row in rows[1:]
    ]


def _export_cell_value(value: object) -> export_module.ExportValue:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _reference_identity_key(
    row: dict[str, export_module.ExportValue],
) -> tuple[object, ...]:
    return (
        row.get("canonical_name"),
        row.get("hp_current"),
        row.get("hp_max"),
        row.get("weight_kg"),
        row.get("iv_attack"),
        row.get("iv_defense"),
        row.get("iv_stamina"),
    )


def test_run_production_export_omits_missing_weight_with_strong_evidence(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "frame.png"
    output_dir = tmp_path / "export"
    Image.new("RGB", (8, 8), "white").save(input_path)

    def visual_scanner(frame: FrameCandidate) -> FrameVisualRecord:
        return FrameVisualRecord(
            frame=frame,
            source_file=frame.source_asset.path.name,
            source_type=frame.source_asset.source_type,
            frame_path=str(frame.frame_path),
            frame_index=frame.frame_index,
            timestamp_s=frame.timestamp_s,
            raw_classification="appraisal",
            signals={"stable_detail_signal": True},
            iv_evidence=_iv_evidence(),
            moves_ocr_box=[0.0, 0.0, 0.0, 0.0],
            motion_sample=None,
        )

    def sequence_scanner(
        sequence: list[FrameVisualRecord],
        settings: ScanSettings,
    ) -> ProductionSequenceScanResult:
        del settings
        visual = sequence[0]
        features = _features(
            "has_CP",
            "has_hp",
            "has_story",
            "has_iv",
            "has_iv_complete",
            "has_tag_chips",
        )
        record = FrameScanRecord(
            source_file=visual.source_file,
            source_type=visual.source_type,
            frame_path=visual.frame_path,
            frame_index=visual.frame_index,
            timestamp_s=visual.timestamp_s,
            classification="detail",
            raw_classification="appraisal",
            features=features,
            values={
                "cp": 123,
                "hp": "10/10",
                "story_text": (
                    "This Bulbasaur was caught on 1/2/2026 around Prague, Czechia."
                ),
                "iv_attack": 10,
                "iv_defense": 11,
                "iv_stamina": 12,
            },
            ocr={},
        )
        return ProductionSequenceScanResult(
            records=[record],
            accepted_fields={
                "cp": 123,
                "hp": "10/10",
                "story": record.values["story_text"],
                "iv": (10, 11, 12),
                "tag": True,
            },
            desired_fields={"cp", "hp", "weight", "story", "iv", "tag"},
            requested_ocr_fields_by_frame={visual.frame_index: ()},
            completed=False,
        )

    report = run_production_export(
        ScanSettings(input_path=input_path, output_dir=output_dir, workers=1),
        visual_scanner=visual_scanner,
        sequence_scanner=sequence_scanner,
    )

    assert not report.rows
    assert any(
        "missing core identity field(s): weight_kg" in warning.message
        for warning in report.warnings
    )

    manifest = json.loads(
        (output_dir / "artifacts" / "export_manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["exported_row_count"] == 0
    assert manifest["unresolved_pokemon_like_sequence_count"] == 1
    diagnostics = [
        json.loads(line)
        for line in (output_dir / "artifacts" / "row_diagnostics.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
        if line
    ]
    assert diagnostics[0]["outcome"] == "rejected"
    assert diagnostics[0]["pokemon_like"] is True
    assert diagnostics[0]["field_values"]["cp"] == 123
    assert diagnostics[0]["field_values"]["hp_current"] == 10
    assert "weight_kg" in diagnostics[0]["skip_reasons"][0]
