from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook  # pylint: disable=import-error

from pogo_storage_mapper.classify import (
    BaseStats,
    ClassifierData,
    InputTable,
    InputValue,
    LegacyMoveEntry,
    OutputRow,
    calculate_cp,
    calculate_stat_product,
    classify_inventory_rows,
    highest_level_under_cp_cap,
    run_inventory_classification,
    run_inventory_validation,
    write_classified_xlsx,
)
from pogo_storage_mapper.metadata import (
    BaseStatsEntry,
    CpMultiplierEntry,
    EvolutionEntry,
    MetadataCatalog,
    SpeciesEntry,
)


def _catalog() -> MetadataCatalog:
    return MetadataCatalog(
        species=(
            SpeciesEntry("bulbasaur", "Bulbasaur", 1),
            SpeciesEntry("ivysaur", "Ivysaur", 2),
            SpeciesEntry("venusaur", "Venusaur", 3),
            SpeciesEntry("pidgey", "Pidgey", 16),
            SpeciesEntry("pidgeotto", "Pidgeotto", 17),
            SpeciesEntry("pidgeot", "Pidgeot", 18),
        ),
        evolutions=(
            EvolutionEntry("bulbasaur", "ivysaur"),
            EvolutionEntry("ivysaur", "venusaur"),
            EvolutionEntry("pidgey", "pidgeotto"),
            EvolutionEntry("pidgeotto", "pidgeot"),
        ),
        base_stats=(
            BaseStatsEntry("bulbasaur", "Normal", 118, 111, 128),
            BaseStatsEntry("ivysaur", "Normal", 151, 143, 155),
            BaseStatsEntry("venusaur", "Normal", 198, 189, 190),
            BaseStatsEntry("pidgey", "Normal", 85, 73, 120),
            BaseStatsEntry("pidgeotto", "Normal", 117, 105, 160),
            BaseStatsEntry("pidgeot", "Normal", 166, 154, 195),
        ),
        cp_multipliers=(
            CpMultiplierEntry(1.0, 0.1),
            CpMultiplierEntry(1.5, 0.2),
            CpMultiplierEntry(2.0, 0.3),
        ),
    )


def _data(
    legacy_moves: tuple[LegacyMoveEntry, ...] = (),
) -> ClassifierData:
    return ClassifierData.from_catalog(_catalog(), legacy_moves=legacy_moves)


def test_cp_formula() -> None:
    stats = BaseStats("test", "Normal", 100, 100, 100)

    assert calculate_cp(stats, 0, 0, 0, 1.0) == 1000


def test_highest_level_under_cp_cap() -> None:
    stats = BaseStats("test", "Normal", 100, 100, 100)
    cpms = (CpMultiplierEntry(1.0, 0.5), CpMultiplierEntry(1.5, 1.0))

    assert highest_level_under_cp_cap(stats, 0, 0, 0, 500, cpms) == (
        1.0,
        0.5,
        250,
    )


def test_stat_product_calculation() -> None:
    stats = BaseStats("test", "Normal", 100, 100, 100)

    assert calculate_stat_product(stats, 0, 0, 0, 1.0) == 1_000_000


def test_evolution_candidate_generation() -> None:
    candidates = _data().evolution_candidates("bulbasaur")

    assert [candidate.species_key for candidate in candidates] == [
        "bulbasaur",
        "ivysaur",
        "venusaur",
    ]


def test_attacker_top3_selection() -> None:
    rows: list[dict[str, InputValue]] = [
        {"species": "Pidgey", "iv_attack": 15, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 14, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 13, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 0, "iv_defense": 0, "iv_stamina": 0},
    ]

    classified = classify_inventory_rows(
        InputTable(["species", "iv_attack", "iv_defense", "iv_stamina"], rows),
        _data(),
    )

    assert [row["attacker_rank_in_family"] for row in classified] == [1, 2, 3, ""]
    assert classified[0]["attacker_as_species"] == "Pidgeot"
    assert classified[3]["attacker_top3"] is False


def test_legacy_warning_review_and_detected_move_keep() -> None:
    legacy = (
        LegacyMoveEntry(
            species_key="bulbasaur",
            evolution_species_key="venusaur",
            move_name="Frenzy Plant",
            move_type="charged",
            note="test",
        ),
    )
    rows: list[dict[str, InputValue]] = [
        {
            "species": "Bulbasaur",
            "iv_attack": 0,
            "iv_defense": 0,
            "iv_stamina": 0,
        },
        {
            "species": "Bulbasaur",
            "iv_attack": 14,
            "iv_defense": 15,
            "iv_stamina": 15,
        },
        {
            "species": "Bulbasaur",
            "iv_attack": 13,
            "iv_defense": 15,
            "iv_stamina": 15,
        },
        {
            "species": "Bulbasaur",
            "iv_attack": 15,
            "iv_defense": 15,
            "iv_stamina": 15,
            "charged_move_name": "Frenzy Plant",
        },
    ]

    classified = classify_inventory_rows(
        InputTable(
            [
                "species",
                "iv_attack",
                "iv_defense",
                "iv_stamina",
                "charged_move_name",
            ],
            rows,
        ),
        _data(legacy),
    )

    assert classified[0]["possible_legacy_elite_move"] is True
    assert classified[0]["recommendation"] == "REVIEW"
    assert classified[3]["has_detected_legacy_move"] is True
    assert classified[3]["recommendation"] == "KEEP"


def test_final_recommendation_let_go_for_non_best_complete_row() -> None:
    rows: list[dict[str, InputValue]] = [
        {"species": "Pidgey", "iv_attack": 15, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 14, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 13, "iv_defense": 15, "iv_stamina": 15},
        {"species": "Pidgey", "iv_attack": 0, "iv_defense": 0, "iv_stamina": 0},
    ]

    classified = classify_inventory_rows(
        InputTable(["species", "iv_attack", "iv_defense", "iv_stamina"], rows),
        _data(),
    )

    assert classified[3]["recommendation"] == "LET-GO"


def test_validate_reports_missing_required_columns(tmp_path: Path) -> None:
    path = tmp_path / "inventory.csv"
    path.write_text("species,iv_attack\nBulbasaur,10\n", encoding="utf-8")

    report = run_inventory_validation(path)

    assert report.errors == [
        "Input must contain iv_defense.",
        "Input must contain iv_stamina.",
    ]


def test_classify_writes_output_folder_artifacts(tmp_path: Path) -> None:
    input_path = tmp_path / "inventory.csv"
    output_dir = tmp_path / "classified"
    with input_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["species", "iv_attack", "iv_defense", "iv_stamina"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "species": "Pidgey",
                "iv_attack": 15,
                "iv_defense": 15,
                "iv_stamina": 15,
            }
        )

    report = run_inventory_classification(
        input_path,
        output_dir,
        catalog=_catalog(),
        legacy_moves=(),
    )

    assert report.csv_path == output_dir / "inventory_classified.csv"
    assert report.xlsx_path == output_dir / "inventory_classified.xlsx"
    assert report.manifest_path == output_dir / "artifacts" / "classify_manifest.json"
    assert report.csv_path.exists()
    assert report.xlsx_path.exists()
    assert report.manifest_path.exists()


def test_xlsx_writer_creates_expected_sheets(tmp_path: Path) -> None:
    rows: list[OutputRow] = [
        {
            "species": "Pidgey",
            "recommendation": "KEEP",
            "pvp_gl_best": True,
            "pvp_ul_best": False,
            "pvp_ll_best": False,
            "attacker_top3": True,
            "possible_legacy_elite_move": False,
        }
    ]
    columns = [
        "species",
        "recommendation",
        "pvp_gl_best",
        "pvp_ul_best",
        "pvp_ll_best",
        "attacker_top3",
        "possible_legacy_elite_move",
    ]
    path = tmp_path / "classified.xlsx"

    write_classified_xlsx(path, columns, rows)

    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "All",
        "KEEP",
        "REVIEW",
        "LET-GO",
        "PvP Great League",
        "PvP Ultra League",
        "PvP Little League",
        "Attacker Top 3",
        "Legacy Warnings",
        "Summary",
    ]
    assert workbook["All"].freeze_panes == "A2"
    assert workbook["PvP Great League"]["A2"].value == "Pidgey"
