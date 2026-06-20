from __future__ import annotations

import csv
import json
import math
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from importlib import resources
from pathlib import Path
from typing import Any, TypeAlias, cast

from openpyxl import Workbook, load_workbook  # pylint: disable=import-error

from pogo_storage_mapper.metadata import (
    BaseStatsEntry,
    CpMultiplierEntry,
    EvolutionEntry,
    MetadataCatalog,
    SpeciesEntry,
    build_family_index,
    build_name_index,
    load_default_metadata_catalog,
    normalize_catalog_name,
    normalize_move_name,
)

InputValue: TypeAlias = str | int | float | bool | None
OutputRow: TypeAlias = dict[str, InputValue]

LEGACY_MOVES_RESOURCE = "data/possible_legacy_elite_moves.csv"

SPECIES_ALIASES = ("species", "species_name", "canonical_name", "species_key")
MOVE_ALIASES = (
    "fast_move",
    "fast_move_name",
    "fast_move_key",
    "charged_move_1",
    "charged_move_1_name",
    "charged_move",
    "charged_move_name",
    "charged_move_key",
    "charged_move_2",
    "charged_move_2_name",
    "second_charged_move",
    "second_charged_move_name",
    "second_charged_move_key",
)
CLASSIFIER_COLUMNS = (
    "family_id",
    "best_final_evolution",
    "attacker_score",
    "attacker_top3",
    "attacker_rank_in_family",
    "attacker_as_species",
    "pvp_ll_best",
    "pvp_ll_as_species",
    "pvp_ll_stat_product",
    "pvp_ll_level",
    "pvp_ll_iv_rank",
    "pvp_gl_best",
    "pvp_gl_as_species",
    "pvp_gl_stat_product",
    "pvp_gl_level",
    "pvp_gl_iv_rank",
    "pvp_ul_best",
    "pvp_ul_as_species",
    "pvp_ul_stat_product",
    "pvp_ul_level",
    "pvp_ul_iv_rank",
    "possible_legacy_elite_move",
    "possible_legacy_elite_moves",
    "has_detected_legacy_move",
    "protection_flags",
    "recommendation",
    "reason",
)
LEAGUES = {
    "ll": ("Little League", 500),
    "gl": ("Great League", 1500),
    "ul": ("Ultra League", 2500),
}
TRUE_TEXT = {"1", "true", "t", "yes", "y"}
FALSE_TEXT = {"0", "false", "f", "no", "n"}


BaseStats = BaseStatsEntry


@dataclass(frozen=True, slots=True)
class EvolutionCandidate:
    species_key: str
    form: str


@dataclass(frozen=True, slots=True)
class LegacyMoveEntry:
    species_key: str | None
    evolution_species_key: str | None
    move_name: str
    move_type: str
    note: str

    @property
    def move_key(self) -> str:
        return normalize_move_name(self.move_name)


@dataclass(slots=True)
class InputTable:
    columns: list[str]
    rows: list[dict[str, InputValue]]


@dataclass(slots=True)
class ValidationReport:
    input_path: Path
    row_count: int
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def summary_line(self) -> str:
        return (
            f"Validated {self.row_count} row(s): "
            f"{len(self.errors)} error(s), {len(self.warnings)} warning(s)."
        )


@dataclass(slots=True)
class ClassifyReport:
    input_path: Path
    output_dir: Path
    csv_path: Path
    xlsx_path: Path
    manifest_path: Path
    row_count: int
    recommendation_counts: dict[str, int]
    warnings: list[str] = field(default_factory=list)

    def summary_line(self) -> str:
        counts = ", ".join(
            f"{name}={self.recommendation_counts.get(name, 0)}"
            for name in ("KEEP", "REVIEW", "LET-GO")
        )
        return f"Classified {self.row_count} row(s): {counts} -> {self.output_dir}"


@dataclass(slots=True)
class ClassifierData:
    species_by_key: dict[str, SpeciesEntry]
    species_name_index: dict[str, SpeciesEntry | None]
    evolutions_by_species: dict[str, list[EvolutionEntry]]
    stats_by_species_form: dict[tuple[str, str], BaseStats]
    stats_forms_by_species: dict[str, list[BaseStats]]
    cp_multipliers: tuple[CpMultiplierEntry, ...]
    family_by_species: dict[str, str]
    legacy_moves: tuple[LegacyMoveEntry, ...]
    iv_rank_cache: dict[tuple[str, str, int], dict[tuple[int, int, int], int]] = field(
        default_factory=dict
    )

    @classmethod
    def from_catalog(
        cls,
        catalog: MetadataCatalog,
        *,
        legacy_moves: tuple[LegacyMoveEntry, ...] = (),
    ) -> ClassifierData:
        species_by_key = {entry.species_key: entry for entry in catalog.species}
        stats_by_species_form: dict[tuple[str, str], BaseStats] = {}
        stats_forms_by_species: dict[str, list[BaseStats]] = {}
        for stats_entry in catalog.base_stats:
            stats = BaseStatsEntry(
                species_key=stats_entry.species_key,
                form=stats_entry.form or "Normal",
                base_attack=stats_entry.base_attack,
                base_defense=stats_entry.base_defense,
                base_stamina=stats_entry.base_stamina,
            )
            stats_by_species_form[(stats_entry.species_key, _form_key(stats.form))] = (
                stats
            )
            stats_forms_by_species.setdefault(stats_entry.species_key, []).append(stats)

        evolutions_by_species: dict[str, list[EvolutionEntry]] = {}
        for evolution_entry in catalog.evolutions:
            evolutions_by_species.setdefault(evolution_entry.species_key, []).append(
                evolution_entry
            )

        return cls(
            species_by_key=species_by_key,
            species_name_index=build_name_index(
                catalog.species,
                lambda entry: (entry.species_name, *entry.aliases, entry.species_key),
                normalize_catalog_name,
            ),
            evolutions_by_species=evolutions_by_species,
            stats_by_species_form=stats_by_species_form,
            stats_forms_by_species=stats_forms_by_species,
            cp_multipliers=tuple(
                sorted(catalog.cp_multipliers, key=lambda entry: entry.level)
            ),
            family_by_species=build_family_index(
                catalog.species,
                catalog.evolutions,
                catalog.base_stats,
            ),
            legacy_moves=legacy_moves,
        )

    @property
    def has_evolution_metadata(self) -> bool:
        return bool(self.evolutions_by_species)

    @property
    def has_pvp_metadata(self) -> bool:
        return bool(self.cp_multipliers and self.stats_by_species_form)

    def species_name(self, species_key: str) -> str:
        entry = self.species_by_key.get(species_key)
        if entry is not None:
            return entry.species_name
        return " ".join(part.capitalize() for part in species_key.split("-"))

    def family_id_for(self, species_key: str) -> str:
        return self.family_by_species.get(species_key, species_key)

    def resolve_species_key(self, row: dict[str, InputValue]) -> str | None:
        species_key = _string_value(_row_value(row, "species_key"))
        if species_key:
            normalized = _key_text(species_key)
            if normalized in self.species_by_key:
                return normalized
        for alias in ("species", "species_name", "canonical_name"):
            raw_value = _string_value(_row_value(row, alias))
            if not raw_value:
                continue
            match = self.species_name_index.get(normalize_catalog_name(raw_value))
            if match is not None:
                return match.species_key
            candidate_key = _key_text(raw_value)
            if candidate_key in self.species_by_key:
                return candidate_key
        return _key_text(species_key) if species_key else None

    def evolution_candidates(
        self, species_key: str, form: str | None = None
    ) -> list[EvolutionCandidate]:
        start = EvolutionCandidate(species_key, _display_form(form))
        candidates: list[EvolutionCandidate] = []
        seen: set[tuple[str, str]] = set()
        queue = [start]
        while queue:
            candidate = queue.pop(0)
            key = (candidate.species_key, _form_key(candidate.form))
            if key in seen:
                continue
            seen.add(key)
            candidates.append(candidate)
            for edge in self.evolutions_by_species.get(candidate.species_key, []):
                if not _edge_applies(edge, candidate.form):
                    continue
                next_form = edge.evolves_to_form or edge.form or candidate.form
                queue.append(
                    EvolutionCandidate(edge.evolves_to_key, _display_form(next_form))
                )
        return candidates

    def final_candidates(
        self, candidates: list[EvolutionCandidate]
    ) -> list[EvolutionCandidate]:
        finals = [
            candidate
            for candidate in candidates
            if not any(
                _edge_applies(edge, candidate.form)
                for edge in self.evolutions_by_species.get(candidate.species_key, [])
            )
        ]
        return finals or candidates[:1]

    def resolve_stats(self, candidate: EvolutionCandidate) -> BaseStats | None:
        exact = self.stats_by_species_form.get(
            (candidate.species_key, _form_key(candidate.form))
        )
        if exact is not None:
            return exact
        normal = self.stats_by_species_form.get(
            (candidate.species_key, _form_key(None))
        )
        if normal is not None:
            return normal
        forms = self.stats_forms_by_species.get(candidate.species_key, [])
        if len(forms) == 1:
            return forms[0]
        return None


@dataclass(slots=True)
class OwnedPokemon:
    index: int
    source_row: dict[str, InputValue]
    output_row: OutputRow
    species_key: str | None
    form: str
    iv_attack: int | None
    iv_defense: int | None
    iv_stamina: int | None
    shadow: bool
    candidates: list[EvolutionCandidate]
    final_candidate: EvolutionCandidate | None
    missing_reasons: list[str] = field(default_factory=list)
    keep_reasons: list[str] = field(default_factory=list)
    review_reasons: list[str] = field(default_factory=list)

    @property
    def ivs(self) -> tuple[int, int, int] | None:
        if self.iv_attack is None or self.iv_defense is None or self.iv_stamina is None:
            return None
        return self.iv_attack, self.iv_defense, self.iv_stamina


@dataclass(frozen=True, slots=True)
class PvpCandidate:
    owned: OwnedPokemon
    species: EvolutionCandidate
    stats: BaseStats
    league_key: str
    league_cap: int
    level: float
    cpm: float
    stat_product: float


def run_inventory_validation(input_path: Path) -> ValidationReport:
    table = read_inventory_table(input_path)
    report = ValidationReport(input_path=input_path, row_count=len(table.rows))
    normalized_columns = {_normalize_column(column) for column in table.columns}
    if not table.rows:
        report.errors.append("Input contains no inventory rows.")
    if not any(alias in normalized_columns for alias in SPECIES_ALIASES):
        report.errors.append("Input must contain a species column or species alias.")
    for column in ("iv_attack", "iv_defense", "iv_stamina"):
        if column not in normalized_columns:
            report.errors.append(f"Input must contain {column}.")
    for row_index, row in enumerate(table.rows, start=2):
        if not any(_present(_row_value(row, alias)) for alias in SPECIES_ALIASES):
            report.warnings.append(f"Row {row_index}: missing species value.")
        for column in ("iv_attack", "iv_defense", "iv_stamina"):
            value = _parse_iv(_row_value(row, column))
            if value is None:
                report.warnings.append(f"Row {row_index}: missing or invalid {column}.")
    return report


def run_inventory_classification(
    input_path: Path,
    output_dir: Path,
    *,
    catalog: MetadataCatalog | None = None,
    legacy_moves: tuple[LegacyMoveEntry, ...] | None = None,
) -> ClassifyReport:
    table = read_inventory_table(input_path)
    validation = run_inventory_validation(input_path)
    if validation.errors:
        msg = "; ".join(validation.errors)
        raise ValueError(msg)

    active_catalog = catalog or load_default_metadata_catalog()
    moves = legacy_moves if legacy_moves is not None else load_default_legacy_moves()
    data = ClassifierData.from_catalog(active_catalog, legacy_moves=moves)
    output_rows = classify_inventory_rows(table, data)
    columns = _output_columns(table.columns)

    output_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir = output_dir / "artifacts"
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    stem = input_path.stem
    csv_path = output_dir / f"{stem}_classified.csv"
    xlsx_path = output_dir / f"{stem}_classified.xlsx"
    manifest_path = artifacts_dir / "classify_manifest.json"

    write_classified_csv(csv_path, columns, output_rows)
    write_classified_xlsx(xlsx_path, columns, output_rows)
    recommendation_counts = Counter(
        str(row.get("recommendation") or "") for row in output_rows
    )
    manifest = {
        "input": str(input_path),
        "row_count": len(output_rows),
        "recommendation_counts": {
            name: recommendation_counts.get(name, 0)
            for name in ("KEEP", "REVIEW", "LET-GO")
        },
        "warning_count": len(validation.warnings),
        "warnings": validation.warnings[:100],
        "metadata": {
            "evolution_entries": len(active_catalog.evolutions),
            "base_stat_entries": len(active_catalog.base_stats),
            "cp_multiplier_entries": len(active_catalog.cp_multipliers),
            "legacy_move_entries": len(moves),
        },
        "artifacts": {
            "classified_csv": str(csv_path),
            "classified_xlsx": str(xlsx_path),
            "classify_manifest": str(manifest_path),
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    _write_json(manifest_path, manifest)
    return ClassifyReport(
        input_path=input_path,
        output_dir=output_dir,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        manifest_path=manifest_path,
        row_count=len(output_rows),
        recommendation_counts=cast(dict[str, int], manifest["recommendation_counts"]),
        warnings=validation.warnings,
    )


def classify_inventory_rows(
    table: InputTable,
    data: ClassifierData,
) -> list[OutputRow]:
    has_shadow_column = any(
        _normalize_column(column) in {"shadow", "is_shadow"} for column in table.columns
    )
    owned_rows = [
        _owned_pokemon(index, row, data, has_shadow_column)
        for index, row in enumerate(table.rows)
    ]
    _apply_attacker_rankings(owned_rows, data, has_shadow_column)
    _apply_pvp_rankings(owned_rows, data)
    for owned in owned_rows:
        _apply_legacy_warnings(owned, data)
        _apply_recommendation(owned)
    return [owned.output_row for owned in owned_rows]


def calculate_cp(
    stats: BaseStats,
    iv_attack: int,
    iv_defense: int,
    iv_stamina: int,
    cpm: float,
) -> int:
    raw_cp = (
        (stats.base_attack + iv_attack)
        * math.sqrt(stats.base_defense + iv_defense)
        * math.sqrt(stats.base_stamina + iv_stamina)
        * cpm
        * cpm
        / 10
    )
    return max(10, math.floor(raw_cp))


# pylint: disable-next=too-many-arguments,too-many-positional-arguments
def highest_level_under_cp_cap(
    stats: BaseStats,
    iv_attack: int,
    iv_defense: int,
    iv_stamina: int,
    cp_cap: int,
    cp_multipliers: tuple[CpMultiplierEntry, ...],
) -> tuple[float, float, int] | None:
    best: tuple[float, float, int] | None = None
    for entry in sorted(cp_multipliers, key=lambda item: item.level):
        cp = calculate_cp(stats, iv_attack, iv_defense, iv_stamina, entry.cpm)
        if cp <= cp_cap:
            best = (entry.level, entry.cpm, cp)
    return best


def calculate_stat_product(
    stats: BaseStats,
    iv_attack: int,
    iv_defense: int,
    iv_stamina: int,
    cpm: float,
) -> float:
    effective_attack = (stats.base_attack + iv_attack) * cpm
    effective_defense = (stats.base_defense + iv_defense) * cpm
    effective_stamina = math.floor((stats.base_stamina + iv_stamina) * cpm)
    return effective_attack * effective_defense * effective_stamina


def read_inventory_table(path: Path) -> InputTable:
    if not path.exists():
        raise ValueError(f"Input file does not exist: {path}")
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _read_csv_table(path)
    if suffix == ".xlsx":
        return _read_xlsx_table(path)
    raise ValueError("Input must be a CSV or XLSX file.")


def write_classified_csv(
    path: Path,
    columns: list[str],
    rows: list[OutputRow],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _csv_cell(row.get(column)) for column in columns})


def write_classified_xlsx(
    path: Path,
    columns: list[str],
    rows: list[OutputRow],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    all_sheet = cast(Any, workbook.active)
    all_sheet.title = "All"
    _write_sheet(all_sheet, columns, rows)

    for recommendation in ("KEEP", "REVIEW", "LET-GO"):
        sheet = workbook.create_sheet(recommendation)
        _write_sheet(
            sheet,
            columns,
            [row for row in rows if row.get("recommendation") == recommendation],
        )

    league_sheets = (
        ("PvP Great League", "pvp_gl_best"),
        ("PvP Ultra League", "pvp_ul_best"),
        ("PvP Little League", "pvp_ll_best"),
    )
    for sheet_name, column in league_sheets:
        sheet = workbook.create_sheet(sheet_name)
        _write_sheet(sheet, columns, [row for row in rows if row.get(column) is True])

    attacker_sheet = workbook.create_sheet("Attacker Top 3")
    _write_sheet(
        attacker_sheet,
        columns,
        [row for row in rows if row.get("attacker_top3") is True],
    )
    legacy_sheet = workbook.create_sheet("Legacy Warnings")
    _write_sheet(
        legacy_sheet,
        columns,
        [row for row in rows if row.get("possible_legacy_elite_move") is True],
    )
    summary_sheet = workbook.create_sheet("Summary")
    _write_summary_sheet(summary_sheet, rows)
    workbook.save(path)


def load_default_legacy_moves() -> tuple[LegacyMoveEntry, ...]:
    try:
        resource = resources.files("pogo_storage_mapper").joinpath(
            LEGACY_MOVES_RESOURCE
        )
        with resource.open("r", encoding="utf-8") as handle:
            return load_legacy_moves(handle)
    except FileNotFoundError:
        return ()


def load_legacy_moves(handle: Any) -> tuple[LegacyMoveEntry, ...]:
    entries: list[LegacyMoveEntry] = []
    reader = csv.DictReader(handle)
    for row in reader:
        species = _key_text(row.get("species"))
        evolution_species = _key_text(row.get("evolution_species"))
        move_name = (row.get("move_name") or "").strip()
        if not move_name:
            continue
        entries.append(
            LegacyMoveEntry(
                species_key=species or None,
                evolution_species_key=evolution_species or None,
                move_name=move_name,
                move_type=(row.get("move_type") or "").strip(),
                note=(row.get("note") or "").strip(),
            )
        )
    return tuple(entries)


def _owned_pokemon(
    index: int,
    row: dict[str, InputValue],
    data: ClassifierData,
    has_shadow_column: bool,
) -> OwnedPokemon:
    output = dict(row)
    for column in CLASSIFIER_COLUMNS:
        output[column] = _default_classifier_value(column)

    species_key = data.resolve_species_key(row)
    form = _display_form(_string_value(_row_value(row, "form")))
    iv_attack = _parse_iv(_row_value(row, "iv_attack"))
    iv_defense = _parse_iv(_row_value(row, "iv_defense"))
    iv_stamina = _parse_iv(_row_value(row, "iv_stamina"))
    shadow = (
        _parse_bool(_row_value(row, "shadow"))
        or _parse_bool(_row_value(row, "is_shadow"))
        if has_shadow_column
        else False
    )

    missing_reasons: list[str] = []
    candidates: list[EvolutionCandidate] = []
    final_candidate = None
    if species_key is None:
        missing_reasons.append("missing species")
    else:
        candidates = data.evolution_candidates(species_key, form)
        final_candidate = _select_attacker_final(data, candidates)
        output["family_id"] = data.family_id_for(species_key)
        if final_candidate is not None:
            output["best_final_evolution"] = data.species_name(
                final_candidate.species_key
            )
    if iv_attack is None or iv_defense is None or iv_stamina is None:
        missing_reasons.append("missing IV triplet")
    else:
        output["attacker_score"] = attacker_score(iv_attack, iv_defense, iv_stamina)
    if not data.has_evolution_metadata:
        missing_reasons.append("missing evolution metadata")
    if not data.has_pvp_metadata:
        missing_reasons.append("missing PvP stats metadata")

    return OwnedPokemon(
        index=index,
        source_row=row,
        output_row=output,
        species_key=species_key,
        form=form,
        iv_attack=iv_attack,
        iv_defense=iv_defense,
        iv_stamina=iv_stamina,
        shadow=shadow,
        candidates=candidates,
        final_candidate=final_candidate,
        missing_reasons=missing_reasons,
    )


def attacker_score(iv_attack: int, iv_defense: int, iv_stamina: int) -> int:
    return 10 * iv_attack + iv_defense + iv_stamina


def _apply_attacker_rankings(
    owned_rows: list[OwnedPokemon],
    data: ClassifierData,
    has_shadow_column: bool,
) -> None:
    groups: dict[tuple[str, str, bool], list[OwnedPokemon]] = {}
    for owned in owned_rows:
        if owned.ivs is None or owned.final_candidate is None:
            continue
        group_key = (
            owned.final_candidate.species_key,
            _form_key(owned.final_candidate.form),
            owned.shadow if has_shadow_column else False,
        )
        groups.setdefault(group_key, []).append(owned)

    for group in groups.values():
        ranked = sorted(
            group,
            key=lambda item: (
                -(cast(int, item.output_row["attacker_score"])),
                -(item.iv_attack or 0),
                -(item.iv_defense or 0),
                -(item.iv_stamina or 0),
                item.index,
            ),
        )
        for rank, owned in enumerate(ranked[:3], start=1):
            final_candidate = owned.final_candidate
            if final_candidate is None:
                continue
            owned.output_row["attacker_top3"] = True
            owned.output_row["attacker_rank_in_family"] = rank
            owned.output_row["attacker_as_species"] = data.species_name(
                final_candidate.species_key
            )
            owned.keep_reasons.append(f"attacker top {rank}")


def _apply_pvp_rankings(
    owned_rows: list[OwnedPokemon],
    data: ClassifierData,
) -> None:
    best_by_family_league: dict[tuple[str, str], PvpCandidate] = {}
    for owned in owned_rows:
        ivs = owned.ivs
        if owned.species_key is None or ivs is None:
            continue
        for species in owned.candidates:
            stats = data.resolve_stats(species)
            if stats is None:
                continue
            for league_key, (_league_name, league_cap) in LEAGUES.items():
                level = highest_level_under_cp_cap(
                    stats, *ivs, league_cap, data.cp_multipliers
                )
                if level is None:
                    continue
                pokemon_level, cpm, _cp = level
                candidate = PvpCandidate(
                    owned=owned,
                    species=EvolutionCandidate(species.species_key, stats.form),
                    stats=stats,
                    league_key=league_key,
                    league_cap=league_cap,
                    level=pokemon_level,
                    cpm=cpm,
                    stat_product=calculate_stat_product(stats, *ivs, cpm),
                )
                group_key = (data.family_id_for(owned.species_key), league_key)
                current = best_by_family_league.get(group_key)
                if current is None or _pvp_sort_key(candidate) > _pvp_sort_key(current):
                    best_by_family_league[group_key] = candidate

    for candidate in best_by_family_league.values():
        owned = candidate.owned
        ivs = owned.ivs
        if ivs is None:
            continue
        prefix = f"pvp_{candidate.league_key}"
        owned.output_row[f"{prefix}_best"] = True
        owned.output_row[f"{prefix}_as_species"] = data.species_name(
            candidate.species.species_key
        )
        owned.output_row[f"{prefix}_stat_product"] = round(candidate.stat_product, 3)
        owned.output_row[f"{prefix}_level"] = candidate.level
        owned.output_row[f"{prefix}_iv_rank"] = _iv_rank(
            data,
            candidate.stats,
            candidate.league_cap,
            *ivs,
        )
        owned.keep_reasons.append(f"best owned PvP {candidate.league_key.upper()}")


def _apply_legacy_warnings(owned: OwnedPokemon, data: ClassifierData) -> None:
    candidate_species = {candidate.species_key for candidate in owned.candidates}
    if owned.species_key is not None:
        candidate_species.add(owned.species_key)
    matching = [
        entry
        for entry in data.legacy_moves
        if (
            entry.species_key in candidate_species
            or entry.evolution_species_key in candidate_species
        )
    ]
    if not matching:
        return

    move_texts = {
        normalize_move_name(value)
        for value in (
            _string_value(_row_value(owned.source_row, alias)) for alias in MOVE_ALIASES
        )
        if value
    }
    legacy_moves = sorted(
        {
            _legacy_move_label(data, entry)
            for entry in matching
            if entry.move_name.strip()
        }
    )
    owned.output_row["possible_legacy_elite_move"] = True
    owned.output_row["possible_legacy_elite_moves"] = ", ".join(legacy_moves)
    owned.review_reasons.append("possible legacy/Elite move")
    if any(entry.move_key in move_texts for entry in matching):
        owned.output_row["has_detected_legacy_move"] = True
        owned.keep_reasons.append("detected legacy move")


def _apply_recommendation(owned: OwnedPokemon) -> None:
    for reason in owned.missing_reasons:
        if reason not in owned.review_reasons:
            owned.review_reasons.append(reason)

    if owned.keep_reasons:
        owned.output_row["recommendation"] = "KEEP"
        reasons = owned.keep_reasons + [
            reason for reason in owned.review_reasons if reason.startswith("possible")
        ]
    elif owned.review_reasons:
        owned.output_row["recommendation"] = "REVIEW"
        reasons = owned.review_reasons
    else:
        owned.output_row["recommendation"] = "LET-GO"
        reasons = ["not a top owned attacker or PvP candidate"]
    owned.output_row["reason"] = "; ".join(dict.fromkeys(reasons))


def _select_attacker_final(
    data: ClassifierData,
    candidates: list[EvolutionCandidate],
) -> EvolutionCandidate | None:
    finals = data.final_candidates(candidates)
    if not finals:
        return None

    def sort_key(candidate: EvolutionCandidate) -> tuple[int, int, int, str]:
        stats = data.resolve_stats(candidate)
        if stats is None:
            return (0, 0, 0, candidate.species_key)
        return (
            stats.base_attack,
            stats.base_attack + stats.base_defense + stats.base_stamina,
            data.species_by_key.get(
                candidate.species_key,
                SpeciesEntry(candidate.species_key, candidate.species_key, 9999),
            ).pokedex_id,
            candidate.species_key,
        )

    return max(finals, key=sort_key)


def _pvp_sort_key(candidate: PvpCandidate) -> tuple[float, float, int]:
    ivs = candidate.owned.ivs or (0, 0, 0)
    return (
        candidate.stat_product,
        candidate.level,
        sum(ivs),
    )


# pylint: disable-next=too-many-arguments,too-many-positional-arguments
def _iv_rank(
    data: ClassifierData,
    stats: BaseStats,
    cp_cap: int,
    iv_attack: int,
    iv_defense: int,
    iv_stamina: int,
) -> int | None:
    cache_key = (stats.species_key, _form_key(stats.form), cp_cap)
    ranking = data.iv_rank_cache.get(cache_key)
    if ranking is None:
        ranked: list[tuple[float, float, tuple[int, int, int]]] = []
        for attack in range(16):
            for defense in range(16):
                for stamina in range(16):
                    level = highest_level_under_cp_cap(
                        stats,
                        attack,
                        defense,
                        stamina,
                        cp_cap,
                        data.cp_multipliers,
                    )
                    if level is None:
                        continue
                    pokemon_level, cpm, _cp = level
                    ranked.append(
                        (
                            calculate_stat_product(
                                stats, attack, defense, stamina, cpm
                            ),
                            pokemon_level,
                            (attack, defense, stamina),
                        )
                    )
        ranked.sort(key=lambda item: (-item[0], -item[1], item[2]))
        ranking = {
            ivs: rank for rank, (_product, _level, ivs) in enumerate(ranked, start=1)
        }
        data.iv_rank_cache[cache_key] = ranking
    return ranking.get((iv_attack, iv_defense, iv_stamina))


def _read_csv_table(path: Path) -> InputTable:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    if not columns:
        raise ValueError("CSV input must contain a header row.")
    return InputTable(columns=columns, rows=cast(list[dict[str, InputValue]], rows))


def _read_xlsx_table(path: Path) -> InputTable:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX input must contain a header row.")
    columns = [str(value).strip() if value is not None else "" for value in rows[0]]
    columns = [column for column in columns if column]
    if not columns:
        raise ValueError("XLSX input must contain a header row.")
    table_rows: list[dict[str, InputValue]] = []
    for raw_row in rows[1:]:
        values = list(raw_row[: len(columns)])
        if not any(_present(value) for value in values):
            continue
        table_rows.append(
            {
                column: _input_cell_value(
                    values[index] if index < len(values) else None
                )
                for index, column in enumerate(columns)
            }
        )
    return InputTable(columns=columns, rows=table_rows)


def _input_cell_value(value: object) -> InputValue:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _write_sheet(sheet: Any, columns: list[str], rows: list[OutputRow]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([_xlsx_cell(row.get(column)) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _set_readable_widths(sheet, columns, rows)


def _write_summary_sheet(sheet: Any, rows: list[OutputRow]) -> None:
    counts = Counter(str(row.get("recommendation") or "") for row in rows)
    sheet.append(["metric", "value"])
    sheet.append(["rows", len(rows)])
    for recommendation in ("KEEP", "REVIEW", "LET-GO"):
        sheet.append([recommendation, counts.get(recommendation, 0)])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14


def _set_readable_widths(
    sheet: Any,
    columns: list[str],
    rows: list[OutputRow],
) -> None:
    for column_index, column in enumerate(columns, start=1):
        values = [column, *(str(row.get(column) or "") for row in rows[:100])]
        width = min(max(len(value) for value in values) + 2, 48)
        sheet.column_dimensions[
            sheet.cell(row=1, column=column_index).column_letter
        ].width = width


def _output_columns(input_columns: list[str]) -> list[str]:
    existing = {_normalize_column(column) for column in input_columns}
    return input_columns + [
        column
        for column in CLASSIFIER_COLUMNS
        if _normalize_column(column) not in existing
    ]


def _default_classifier_value(column: str) -> InputValue:
    if column.endswith("_best") or column in {
        "attacker_top3",
        "possible_legacy_elite_move",
        "has_detected_legacy_move",
    }:
        return False
    return ""


def _row_value(row: dict[str, InputValue], alias: str) -> InputValue:
    normalized_alias = _normalize_column(alias)
    for key, value in row.items():
        if _normalize_column(key) == normalized_alias:
            return value
    return None


def _string_value(value: InputValue) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _parse_iv(value: InputValue) -> int | None:
    text = _string_value(value)
    if not text:
        return None
    try:
        parsed = int(float(text))
    except ValueError:
        return None
    return parsed if 0 <= parsed <= 15 else None


def _parse_bool(value: InputValue) -> bool:
    if isinstance(value, bool):
        return value
    text = _string_value(value).casefold()
    if text in TRUE_TEXT:
        return True
    if text in FALSE_TEXT:
        return False
    return False


def _present(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _normalize_column(column: str) -> str:
    return "_".join(
        "".join(
            character if character.isalnum() else " " for character in column.casefold()
        ).split()
    )


def _form_key(form: str | None) -> str:
    return normalize_catalog_name(form or "Normal")


def _display_form(form: str | None) -> str:
    return form.strip() if isinstance(form, str) and form.strip() else "Normal"


def _key_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().casefold().replace("_", "-")
    characters = [
        character if character.isalnum() else "-" for character in text if character
    ]
    return "-".join(part for part in "".join(characters).split("-") if part)


def _edge_applies(edge: EvolutionEntry, form: str) -> bool:
    return edge.form is None or _form_key(edge.form) == _form_key(form)


def _legacy_move_label(data: ClassifierData, entry: LegacyMoveEntry) -> str:
    species_key = entry.evolution_species_key or entry.species_key or ""
    species_name = data.species_name(species_key) if species_key else "Unknown"
    note = f" ({entry.note})" if entry.note else ""
    return f"{species_name}: {entry.move_name}{note}"


def _csv_cell(value: InputValue) -> InputValue | str:
    return "" if value is None else value


def _xlsx_cell(value: InputValue) -> InputValue | str:
    return "" if value is None else value


def _write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
